"""Functions for shaping and formatting spreadsheets."""

import math
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from typeguard import typechecked

from bfb_delivery.lib.constants import (
    BOX_TYPE_COLOR_MAP,
    COLUMN_NAME_MAP,
    COMBINED_ROUTES_COLUMNS,
    FILE_DATE_FORMAT,
    FORMATTED_ROUTES_COLUMNS,
    MANIFEST_DATE_FORMAT,
    NOTES_COLUMN_WIDTH,
    PROTEIN_BOX_TYPES,
    SPLIT_ROUTE_COLUMNS,
    BoxType,
    CellColors,
    Columns,
)
from bfb_delivery.lib.formatting.data_cleaning import (
    format_and_validate_data,
    format_column_names,
)
from bfb_delivery.utils import get_phone_number, map_columns

# Silences warning for in-place operations on copied df slices.
pd.options.mode.copy_on_write = True


# TODO: When wrapping in final function, start calling it "make_manifest" or similar.
# TODO: There's got to be a way to set the docstring as a constant.
# TODO: Use Pandera.
# TODO: Switch to or allow CSVs instead of Excel files.
@typechecked
def split_chunked_route(
    input_path: Path | str, output_dir: Path | str, output_filename: str, n_books: int
) -> list[Path]:
    """See public docstring: :py:func:`bfb_delivery.api.public.split_chunked_route`."""
    if n_books <= 0:
        raise ValueError("n_books must be greater than 0.")
    # TODO: Make this accept input_path only as Path? Or only as str to simplify?
    input_path = Path(input_path)

    chunked_sheet: pd.DataFrame = pd.read_excel(input_path)
    chunked_sheet.columns = format_column_names(columns=chunked_sheet.columns.to_list())
    map_columns(df=chunked_sheet, column_name_map=COLUMN_NAME_MAP, invert_map=False)
    format_and_validate_data(df=chunked_sheet, columns=SPLIT_ROUTE_COLUMNS + [Columns.DRIVER])
    chunked_sheet.sort_values(by=[Columns.DRIVER, Columns.STOP_NO], inplace=True)
    # TODO: Validate columns? (Use Pandera?)

    drivers = sorted(list(chunked_sheet[Columns.DRIVER].unique()))
    driver_count = len(drivers)
    if driver_count < n_books:
        raise ValueError(
            "n_books must be less than or equal to the number of drivers: "
            f"driver_count: {driver_count}, n_books: {n_books}."
        )

    output_dir = Path(output_dir) if output_dir else Path(input_path).parent
    base_output_filename = (
        f"split_workbook_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
        if output_filename == ""
        else output_filename
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    split_workbook_paths: list[Path] = []
    driver_sets = [drivers[i::n_books] for i in range(n_books)]
    for i, driver_set in enumerate(driver_sets):
        i_file_name = f"{base_output_filename.split('.')[0]}_{i + 1}.xlsx"
        split_workbook_path: Path = output_dir / i_file_name
        split_workbook_paths.append(split_workbook_path)

        with pd.ExcelWriter(split_workbook_path) as writer:
            driver_set_df = chunked_sheet[chunked_sheet[Columns.DRIVER].isin(driver_set)]
            driver_set_df.sort_values(by=[Columns.DRIVER, Columns.STOP_NO], inplace=True)
            for driver_name, data in driver_set_df.groupby(Columns.DRIVER):
                data[SPLIT_ROUTE_COLUMNS].to_excel(
                    writer, sheet_name=str(driver_name), index=False
                )

    split_workbook_paths = [path.resolve() for path in split_workbook_paths]

    return split_workbook_paths


@typechecked
def create_manifests(
    input_dir: Path | str, output_dir: Path | str, output_filename: str, date: str
) -> Path:
    """See public docstring for :py:func:`bfb_delivery.api.public.create_manifests`."""
    output_filename = (
        f"final_manifests_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
        if output_filename == ""
        else output_filename
    )

    combined_route_workbook_path = combine_route_tables(
        input_dir=input_dir, output_dir=output_dir, output_filename=""
    )

    formatted_manifest_path = format_combined_routes(
        input_path=combined_route_workbook_path,
        output_dir=output_dir,
        output_filename=output_filename,
        date=date,
    )

    return formatted_manifest_path


@typechecked
def combine_route_tables(
    input_dir: Path | str, output_dir: Path | str, output_filename: str
) -> Path:
    """See public docstring: :py:func:`bfb_delivery.api.public.combine_route_tables`."""
    input_dir = Path(input_dir)
    paths = list(input_dir.glob("*.csv"))

    output_dir = Path(output_dir) if output_dir else paths[0].parent
    output_filename = (
        f"combined_routes_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
        if output_filename == ""
        else output_filename
    )
    output_path = output_dir / output_filename
    output_dir.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path) as writer:
        for path in sorted(paths):
            route_df = pd.read_csv(path)
            map_columns(df=route_df, column_name_map=COLUMN_NAME_MAP, invert_map=True)
            route_df.sort_values(by=[Columns.STOP_NO], inplace=True)
            driver_name = path.stem
            route_df[COMBINED_ROUTES_COLUMNS].to_excel(
                writer, sheet_name=driver_name, index=False
            )

    return output_path.resolve()


@typechecked
def format_combined_routes(
    input_path: Path | str, output_dir: Path | str, output_filename: str, date: str
) -> Path:
    """See public docstring: :py:func:`bfb_delivery.api.public.format_combined_routes`."""
    input_path = Path(input_path)
    output_dir = Path(output_dir) if output_dir else input_path.parent
    output_filename = (
        f"formatted_routes_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
        if output_filename == ""
        else output_filename
    )
    output_path = Path(output_dir) / output_filename
    friday = datetime.now() + pd.DateOffset(weekday=4)
    date = date if date else friday.strftime(MANIFEST_DATE_FORMAT)

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb["Sheet"])
    with pd.ExcelFile(input_path) as xls:
        for sheet_idx, sheet_name in enumerate(sorted(xls.sheet_names)):

            driver_name = str(sheet_name)
            new_sheet_name = f"{date} {driver_name}"
            route_df = pd.read_excel(xls, driver_name)

            # TODO: Use Pandera?
            route_df.columns = format_column_names(columns=route_df.columns.to_list())
            format_and_validate_data(df=route_df, columns=COMBINED_ROUTES_COLUMNS)

            # TODO: Order by apartment number, and redo stop numbers?
            # May need to postpone this.
            # Or, for now, just do it if the apartments are already in contiguous stops.
            # Or if discontinuous, just regroup and bump the following stops.
            # Also, may not make the most sense in order of apt number. Ask team.
            route_df.sort_values(by=[Columns.STOP_NO], inplace=True)

            agg_dict = _aggregate_route_data(df=route_df)
            # TODO: !! What happens when there are more than one order for a stop? Two rows?
            # (Since order count column is dropped in manifest)
            # Oh wait, they're all 1s, so is that just a way for them to count them with sum?
            # If that's so, ignore it or validate always a 1?

            ws = wb.create_sheet(title=new_sheet_name, index=sheet_idx)
            _make_manifest_sheet(
                ws=ws,
                agg_dict=agg_dict,
                route_df=route_df,
                date=date,
                driver_name=driver_name,
            )

    # Can check cell values, though. (Maye read dataframe from start row?)
    wb.save(output_path)

    return output_path.resolve()


@typechecked
def _aggregate_route_data(df: pd.DataFrame) -> dict:
    """Aggregate data for a single route.

    Args:
        df: The route data to aggregate.

    Returns:
        Dictionary of aggregated data.
    """
    df = df.copy()

    df[Columns.BOX_TYPE] = df[Columns.BOX_TYPE].str.upper().str.strip()
    box_types = df[Columns.BOX_TYPE].unique()
    extra_box_types = set(box_types) - set(BoxType)
    if extra_box_types:
        raise ValueError(f"Invalid box type in route data: {extra_box_types}")

    agg_dict = {
        "box_counts": df.groupby(Columns.BOX_TYPE)[Columns.ORDER_COUNT].sum().to_dict(),
        "total_box_count": df[Columns.ORDER_COUNT].sum(),
        "protein_box_count": df[df[Columns.BOX_TYPE].isin(PROTEIN_BOX_TYPES)][
            Columns.ORDER_COUNT
        ].sum(),
        "neighborhoods": df[Columns.NEIGHBORHOOD].unique().tolist(),
    }

    for box_type in BoxType:
        if box_type.value not in agg_dict["box_counts"]:
            agg_dict["box_counts"][box_type] = 0

    return agg_dict


@typechecked
def _make_manifest_sheet(
    ws: Worksheet, agg_dict: dict, route_df: pd.DataFrame, date: str, driver_name: str
) -> None:
    """Create a manifest sheet."""
    _add_header_row(ws=ws)
    neighborhoods_row_number = _add_aggregate_block(
        ws=ws, agg_dict=agg_dict, date=date, driver_name=driver_name
    )
    df_start_row = _write_data_to_sheet(ws=ws, df=route_df)
    _auto_adjust_column_widths(ws=ws, df_start_row=df_start_row)
    _word_wrap_notes_column(ws=ws)
    _merge_and_wrap_neighborhoods(ws=ws, neighborhoods_row_number=neighborhoods_row_number)

    # TODO: Set print_area (Use calculate_dimensions)
    # TODO: set_printer_settings(paper_size, orientation)


@typechecked
def _add_header_row(ws: Worksheet) -> None:
    """Append a reusable formatted row to the worksheet."""
    font = Font(bold=True)
    alignment_left = Alignment(horizontal="left")
    alignment_right = Alignment(horizontal="right")
    fill = PatternFill(
        start_color=CellColors.HEADER, end_color=CellColors.HEADER, fill_type="solid"
    )

    driver_support_phone = get_phone_number("driver_support")
    recipient_support_phone = get_phone_number("recipient_support")
    formatted_row = [
        {
            "value": f"DRIVER SUPPORT: {driver_support_phone}",
            "font": font,
            "alignment": alignment_left,
            "fill": fill,
        },
        {"value": "", "font": font, "alignment": None, "fill": fill},
        {"value": "", "font": font, "alignment": None, "fill": fill},
        {
            "value": f"RECIPIENT SUPPORT: {recipient_support_phone}",
            "font": font,
            "alignment": alignment_right,
            "fill": fill,
        },
        {"value": "", "font": font, "alignment": None, "fill": fill},
        {
            "value": "PLEASE SHRED MANIFEST AFTER COMPLETING ROUTE.",
            "font": font,
            "alignment": alignment_right,
            "fill": fill,
        },
    ]

    for col_idx, col_data in enumerate(formatted_row, start=1):
        cell_value = col_data["value"] if isinstance(col_data["value"], str) else ""
        cell = ws.cell(row=1, column=col_idx, value=cell_value)
        if col_data["font"]:
            if isinstance(col_data["font"], Font):
                cell.font = col_data["font"]
        if col_data["alignment"]:
            if isinstance(col_data["alignment"], Alignment):
                cell.alignment = col_data["alignment"]
        if col_data["fill"]:
            if isinstance(col_data["fill"], PatternFill):
                cell.fill = col_data["fill"]

    return


@typechecked
def _add_aggregate_block(ws: Worksheet, agg_dict: dict, date: str, driver_name: str) -> int:
    """Append left and right aggregation blocks to the worksheet row by row."""
    # TODO: Yeah, let's use an enum for box types since the manifest is a contract.
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment_left = Alignment(horizontal="left")
    alignment_right = Alignment(horizontal="right")
    bold_font = Font(bold=True)

    left_block = [
        [{"value": None}],
        [{"value": f"Date: {date}"}],
        [{"value": None}],
        [{"value": f"Driver: {driver_name}"}],
        [{"value": None}],
        [{"value": f"Neighborhoods: {', '.join(agg_dict['neighborhoods'])}"}],
        [{"value": None}],
    ]

    right_block = [
        [{"value": None, "fill": None, "border": None}],
        [
            {
                "value": "BASIC",
                "fill": PatternFill(
                    start_color=CellColors.BASIC,
                    end_color=CellColors.BASIC,
                    fill_type="solid",
                ),
                "border": thin_border,
            },
            {
                "value": agg_dict["box_counts"].get("BASIC", 0),
                "fill": None,
                "border": thin_border,
            },
        ],
        [
            {
                "value": "GF",
                "fill": PatternFill(
                    start_color=CellColors.GF, end_color=CellColors.GF, fill_type="solid"
                ),
                "border": thin_border,
            },
            {
                "value": agg_dict["box_counts"].get("GF", 0),
                "fill": None,
                "border": thin_border,
            },
        ],
        [
            {
                "value": "LA",
                "fill": PatternFill(
                    start_color=CellColors.LA, end_color=CellColors.LA, fill_type="solid"
                ),
                "border": thin_border,
            },
            {
                "value": agg_dict["box_counts"].get("LA", 0),
                "fill": None,
                "border": thin_border,
            },
        ],
        [
            {
                "value": "VEGAN",
                "fill": PatternFill(
                    start_color=CellColors.VEGAN,
                    end_color=CellColors.VEGAN,
                    fill_type="solid",
                ),
                "border": thin_border,
            },
            {
                "value": agg_dict["box_counts"].get("VEGAN", 0),
                "fill": None,
                "border": thin_border,
            },
        ],
        [
            {"value": "TOTAL BOX COUNT=", "fill": None, "border": None},
            {"value": agg_dict["total_box_count"], "fill": None, "border": None},
        ],
        [
            {"value": "PROTEIN COUNT=", "fill": None, "border": None},
            {"value": agg_dict["protein_box_count"], "fill": None, "border": None},
        ],
    ]

    start_row = ws.max_row + 1
    neighborhoods_row_number = 0
    for i, (left_row, right_row) in enumerate(
        zip(left_block, right_block, strict=True), start=start_row
    ):
        for col_idx, cell_definition in enumerate(left_row, start=1):
            cell = ws.cell(row=i, column=col_idx, value=cell_definition["value"])
            cell.font = bold_font
            cell.alignment = alignment_left
            if cell_definition["value"] and cell_definition["value"].startswith(
                "Neighborhoods"
            ):
                neighborhoods_row_number = i

        for col_idx, cell_definition in enumerate(right_row, start=5):
            cell = ws.cell(row=i, column=col_idx, value=cell_definition["value"])
            cell.font = bold_font
            cell.alignment = alignment_right
            if isinstance(cell_definition["fill"], PatternFill):
                cell.fill = cell_definition["fill"]
            if isinstance(cell_definition["border"], Border):
                cell.border = cell_definition["border"]

    return neighborhoods_row_number


@typechecked
def _write_data_to_sheet(ws: Worksheet, df: pd.DataFrame) -> int:
    """Write and format the dataframe itself."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_font = Font(bold=True)

    box_type_col_idx = df.columns.get_loc(Columns.BOX_TYPE)

    start_row = ws.max_row + 1
    for r_idx, row in enumerate(
        dataframe_to_rows(df[FORMATTED_ROUTES_COLUMNS], index=False, header=True),
        start=start_row,
    ):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == start_row:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left")

            if c_idx == box_type_col_idx and r_idx > start_row:
                box_type = str(value)
                fill_color = BOX_TYPE_COLOR_MAP.get(box_type)
                if fill_color:
                    cell.fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )

    return start_row


@typechecked
def _auto_adjust_column_widths(ws: Worksheet, df_start_row: int) -> None:
    """Auto-adjust column widths to fit the dataframe."""
    for col in ws.columns:
        max_length = 0

        col_letter = col[0].column_letter
        padding_scalar = 0.9 if col_letter == "C" else 1  # C is address column.
        for cell in col:
            if cell.row >= df_start_row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)) * padding_scalar)
                except Exception as e:
                    warnings.warn(f"Error while adjusting column widths: {e}", stacklevel=2)

        ws.column_dimensions[col_letter].width = max(8, round(max_length))

    return


@typechecked
def _word_wrap_notes_column(ws: Worksheet) -> None:
    """Word wrap the notes column, and set width."""
    end_row = ws.max_row
    col_letter = "E"
    ws.column_dimensions[col_letter].width = NOTES_COLUMN_WIDTH
    for row in ws[f"{col_letter}10:{col_letter}{end_row}"]:
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    return


@typechecked
def _merge_and_wrap_neighborhoods(ws: Worksheet, neighborhoods_row_number: int) -> None:
    """Merge the neighborhoods cell and wrap the text."""
    start_col = 1
    end_col = 3
    ws.merge_cells(
        start_row=neighborhoods_row_number,
        start_column=start_col,
        end_row=neighborhoods_row_number,
        end_column=end_col,
    )
    cell = ws.cell(row=neighborhoods_row_number, column=start_col)
    cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")

    # Merged cells don't adjust height automatically, so we need to estimate it.
    if cell.value:
        merged_width = sum(
            ws.column_dimensions[col[0].column_letter].width
            for col in ws.iter_cols(min_col=start_col, max_col=end_col)
        )
        char_width = 1.2
        lines = 0
        for line in str(cell.value).split():
            line_length = len(line) * char_width
            lines += line_length / merged_width

        ws.row_dimensions[neighborhoods_row_number].height = max(15, math.ceil(lines) * 15)

    return
