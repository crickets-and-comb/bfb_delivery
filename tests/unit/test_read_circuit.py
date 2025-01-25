"""Tests for read_circuit module."""

import copy
import json
import re
from collections.abc import Iterator
from contextlib import AbstractContextManager, nullcontext
from datetime import datetime
from pathlib import Path
from typing import Any, Final
from unittest.mock import Mock, patch

import numpy as np
import pandas as pd
import pandera as pa
import pytest
import requests
from click.testing import CliRunner
from openpyxl import Workbook, load_workbook
from pandera.errors import SchemaError
from pandera.typing import Series
from pydantic_core import ValidationError
from typeguard import typechecked

from bfb_delivery import create_manifests_from_circuit
from bfb_delivery.cli import (
    create_manifests_from_circuit as create_manifests_from_circuit_cli,
)
from bfb_delivery.lib.constants import (
    ALL_HHS_DRIVER,
    BOX_TYPE_COLOR_MAP,
    FILE_DATE_FORMAT,
    FORMATTED_ROUTES_COLUMNS,
    NOTES_COLUMN_WIDTH,
    BoxType,
    CellColors,
    CircuitColumns,
    Columns,
    IntermediateColumns,
)
from bfb_delivery.lib.dispatch.read_circuit import (
    _get_raw_plans,
    _get_raw_stops,
    _get_responses,
    _make_plans_df,
    _transform_routes_df,
    _write_routes_dfs,
)
from bfb_delivery.lib.formatting.data_cleaning import (
    _format_and_validate_box_type,
    _format_and_validate_name,
    _format_and_validate_phone,
)
from bfb_delivery.lib.formatting.sheet_shaping import _aggregate_route_data
from bfb_delivery.lib.formatting.utils import get_extra_notes
from bfb_delivery.lib.schema import CircuitRoutesTransformInFromDict
from bfb_delivery.lib.schema.checks import at_least_one_in_group, one_to_one

TEST_START_DATE: Final[str] = "2025-01-17"
MANIFEST_DATE: Final[str] = "1.17"


@pytest.fixture()
@typechecked
def mock_plan_responses() -> (
    list[dict[str, str | list[dict[str, str | dict[str, int]] | None]]]
):
    """Return a list of plan responses, as from _get_plan_responses."""
    with open("tests/unit/fixtures/plan_responses.json") as f:
        return json.load(f)


@pytest.fixture()
@typechecked
def mock_get_plan_responses(
    mock_plan_responses: list[dict[str, str | list[dict[str, str | dict[str, int]] | None]]]
) -> Iterator[None]:
    """Mock _get_plan_responses."""
    with patch(
        "bfb_delivery.lib.dispatch.read_circuit._get_plan_responses",
        return_value=mock_plan_responses,
    ):
        yield


@pytest.fixture()
def mock_os_getcwd(tmp_path: Path) -> Iterator[str]:
    """Mock os.getcwd within the read_circuit module."""
    return_value = str(tmp_path)
    with patch("bfb_delivery.lib.dispatch.read_circuit._getcwd", return_value=return_value):
        yield return_value


@pytest.mark.usefixtures("mock_get_plan_responses", "mock_os_getcwd")
class TestCreateManifestsFromCircuit:
    """Test create_manifests_from_circuit function."""

    @pytest.fixture()
    @typechecked
    def mock_stops_responses_all_hhs_false(
        self,
    ) -> list[
        list[
            dict[
                str,
                str
                | list[
                    dict[
                        str,
                        str
                        | int
                        | dict[
                            str, str | int | dict[str, str | int | list[str] | None] | None
                        ]
                        | None,
                    ]
                ]
                | None,
            ]
        ]
    ]:
        """Return a list of stops responses, as from _get_raw_stops_list."""
        with open("tests/unit/fixtures/stops_responses.json") as f:
            return json.load(f)

    @pytest.fixture()
    @typechecked
    def mock_stops_responses_all_hhs_true(
        self,
    ) -> list[
        list[
            dict[
                str,
                str
                | list[
                    dict[
                        str,
                        str
                        | int
                        | dict[
                            str, str | int | dict[str, str | int | list[str] | None] | None
                        ]
                        | None,
                    ]
                ]
                | None,
            ]
        ]
    ]:
        """Return a list of stops responses, as from _get_raw_stops_list."""
        with open("tests/unit/fixtures/stops_responses_all_hhs.json") as f:
            return json.load(f)

    @pytest.fixture()
    @typechecked
    def mock_driver_sheet_names_all_hhs_false(
        self,
        mock_plan_responses: list[
            dict[
                str, str | list[dict[str, str | list[str | dict[str, str]] | dict[str, int]]]
            ]
        ],
    ) -> list[str]:
        """Return a list of driver sheet names."""
        driver_sheet_names = []
        for page_dict in mock_plan_responses:
            for plan_dict in page_dict["plans"]:
                if (
                    isinstance(plan_dict, dict)  # To satisisfy pytype.
                    and isinstance(plan_dict["title"], str)  # To satisisfy pytype.
                    and ALL_HHS_DRIVER not in plan_dict["title"]
                ):
                    driver_sheet_names.append(plan_dict["title"])

        return driver_sheet_names

    @pytest.fixture()
    @typechecked
    def mock_driver_names_all_hhs_false(
        self, mock_driver_sheet_names_all_hhs_false: list[str]
    ) -> list[str]:
        """Return a list of driver names."""
        return [
            " ".join(sheet_name.split(" ")[1:])
            for sheet_name in mock_driver_sheet_names_all_hhs_false
        ]

    @pytest.fixture()
    @typechecked
    def mock_driver_sheet_names_all_hhs_true(
        self,
        mock_plan_responses: list[
            dict[
                str, str | list[dict[str, str | list[str | dict[str, str]] | dict[str, int]]]
            ]
        ],
    ) -> list[str]:
        """Return a list of driver sheet names."""
        driver_sheet_names = []
        for page_dict in mock_plan_responses:
            for plan_dict in page_dict["plans"]:
                if (
                    isinstance(plan_dict, dict)  # To satisfy pytype.
                    and isinstance(plan_dict["title"], str)  # To satisfy pytype.
                    and ALL_HHS_DRIVER in plan_dict["title"]
                ):
                    driver_sheet_names.append(plan_dict["title"])

        return driver_sheet_names

    @pytest.mark.parametrize("circuit_output_dir", ["dummy_circuit_output", ""])
    @pytest.mark.parametrize(
        "all_HHs, mock_stops_responses_fixture",
        [
            (True, "mock_stops_responses_all_hhs_true"),
            (False, "mock_stops_responses_all_hhs_false"),
        ],
    )
    @pytest.mark.parametrize("verbose", [True, False])
    @pytest.mark.parametrize("test_cli", [False, True])
    def test_set_output_dir(
        self,
        circuit_output_dir: str,
        all_HHs: bool,
        mock_stops_responses_fixture: str,
        verbose: bool,
        test_cli: bool,
        mock_os_getcwd: str,
        tmp_path: Path,
        request: pytest.FixtureRequest,
    ) -> None:
        """Test that the output directory can be set."""
        stops_response_data = request.getfixturevalue(mock_stops_responses_fixture)

        output_dir = str(tmp_path / "dummy_output_dir")
        expected_output_filename = (
            f"final_manifests_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
        )
        expected_output_path = Path(output_dir) / expected_output_filename

        circuit_output_dir = (
            str(tmp_path / circuit_output_dir) if circuit_output_dir else circuit_output_dir
        )
        circuit_sub_dir = "routes_" + TEST_START_DATE
        expected_circuit_output_dir = (
            Path(circuit_output_dir) / circuit_sub_dir
            if circuit_output_dir
            else Path(mock_os_getcwd) / circuit_sub_dir
        )

        Path(expected_circuit_output_dir).mkdir(parents=True, exist_ok=True)
        with open(f"{expected_circuit_output_dir}/dummy_file.txt", "w") as f:
            f.write("Dummy file. The function should remove this file.")

        with patch(
            "bfb_delivery.lib.dispatch.read_circuit._get_raw_stops_list",
            return_value=stops_response_data,
        ):
            if test_cli:
                cli_runner = CliRunner()
                arg_list = [
                    "--start_date",
                    TEST_START_DATE,
                    "--output_dir",
                    output_dir,
                    "--circuit_output_dir",
                    circuit_output_dir,
                ]
                if all_HHs:
                    arg_list.append("--all_hhs")
                if verbose:
                    arg_list.append("--verbose")
                result = cli_runner.invoke(create_manifests_from_circuit_cli.main, arg_list)
                assert result.exit_code == 0
                output_path, new_circuit_output_dir = (
                    result.stdout_bytes.decode("utf-8").strip().split("\n")
                )
            else:
                output_path, new_circuit_output_dir = create_manifests_from_circuit(
                    start_date=TEST_START_DATE,
                    output_dir=output_dir,
                    circuit_output_dir=circuit_output_dir,
                    all_HHs=all_HHs,
                    verbose=verbose,
                )

        assert str(new_circuit_output_dir) == str(expected_circuit_output_dir)
        assert str(output_path) == str(expected_output_path)
        assert expected_output_path.exists()

    @pytest.mark.parametrize(
        "all_HHs, mock_stops_responses_fixture, mock_driver_sheet_names_fixture",
        [
            (
                True,
                "mock_stops_responses_all_hhs_true",
                "mock_driver_sheet_names_all_hhs_true",
            ),
            (
                False,
                "mock_stops_responses_all_hhs_false",
                "mock_driver_sheet_names_all_hhs_false",
            ),
        ],
    )
    def test_all_drivers_have_a_sheet(
        self,
        tmp_path: Path,
        all_HHs: bool,
        mock_stops_responses_fixture: str,
        mock_driver_sheet_names_fixture: str,
        request: pytest.FixtureRequest,
    ) -> None:
        """Test that all drivers have a sheet in the formatted workbook. And date works."""
        stops_response_data = request.getfixturevalue(mock_stops_responses_fixture)
        driver_sheet_names = request.getfixturevalue(mock_driver_sheet_names_fixture)
        with patch(
            "bfb_delivery.lib.dispatch.read_circuit._get_raw_stops_list",
            return_value=stops_response_data,
        ):
            output_path, circuit_output_dir = create_manifests_from_circuit(
                start_date=TEST_START_DATE, output_dir=str(tmp_path), all_HHs=all_HHs
            )
        workbook = pd.ExcelFile(output_path)
        assert sorted(list(workbook.sheet_names)) == sorted(driver_sheet_names)
        assert sorted([path.name for path in list(circuit_output_dir.glob("*"))]) == sorted(
            [f"{sheet_name}.csv" for sheet_name in driver_sheet_names]
        )


@pytest.fixture(scope="class")
@typechecked
def mock_plan_responses_basic() -> (
    list[dict[str, str | list[dict[str, str | dict[str, int]] | None]]]
):
    """Return a list of plan responses, as from _get_plan_responses."""
    with open("tests/unit/fixtures/plan_responses.json") as f:
        return json.load(f)


@pytest.fixture(scope="class")
@typechecked
def mock_get_plan_responses_basic(
    mock_plan_responses_basic: list[
        dict[str, str | list[dict[str, str | dict[str, int]] | None]]
    ]
) -> Iterator[None]:
    """Mock _get_plan_responses."""
    with patch(
        "bfb_delivery.lib.dispatch.read_circuit._get_plan_responses",
        return_value=mock_plan_responses_basic,
    ):
        yield


@pytest.fixture(scope="class")
def mock_os_getcwd_class_scope(tmp_path_factory: pytest.TempPathFactory) -> Iterator[str]:
    """Mock os.getcwd within the read_circuit module."""
    return_value = str(tmp_path_factory.mktemp("cwd"))
    with patch("bfb_delivery.lib.dispatch.read_circuit._getcwd", return_value=return_value):
        yield return_value


@pytest.mark.usefixtures("mock_get_plan_responses_basic", "mock_os_getcwd_class_scope")
class TestCreateManifestsFromCircuitClassScoped:
    """Test create_manifests_from_circuit function."""

    @pytest.fixture(scope="class")
    @typechecked
    def mock_driver_sheet_names_basic(
        self,
        mock_plan_responses_basic: list[
            dict[
                str, str | list[dict[str, str | list[str | dict[str, str]] | dict[str, int]]]
            ]
        ],
    ) -> list[str]:
        """Return a list of driver sheet names."""
        driver_sheet_names = []
        for page_dict in mock_plan_responses_basic:
            for plan_dict in page_dict["plans"]:
                if (
                    isinstance(plan_dict, dict)  # To satisisfy pytype.
                    and isinstance(plan_dict["title"], str)  # To satisisfy pytype.
                    and ALL_HHS_DRIVER not in plan_dict["title"]
                ):
                    driver_sheet_names.append(plan_dict["title"])

        return driver_sheet_names

    @pytest.fixture(scope="class")
    @typechecked
    def mock_driver_names_basic(self, mock_driver_sheet_names_basic: list[str]) -> list[str]:
        """Return a list of driver names."""
        return [
            " ".join(sheet_name.split(" ")[1:])
            for sheet_name in mock_driver_sheet_names_basic
        ]

    @pytest.fixture(scope="class")
    @typechecked
    def mock_basic_stops_responses(
        self,
    ) -> list[
        list[
            dict[
                str,
                str
                | list[
                    dict[
                        str,
                        str
                        | int
                        | dict[
                            str, str | int | dict[str, str | int | list[str] | None] | None
                        ]
                        | None,
                    ]
                ]
                | None,
            ]
        ]
    ]:
        """Return a list of stops responses, as from _get_raw_stops_list."""
        with open("tests/unit/fixtures/stops_responses.json") as f:
            return json.load(f)

    @pytest.fixture(scope="class")
    def outputs(
        self, mock_basic_stops_responses: list, tmp_path_factory: pytest.TempPathFactory
    ) -> tuple[Path, Path]:
        """Create a basic manifest scoped to class for reuse."""
        output_dir = str(tmp_path_factory.mktemp("output"))
        with patch(
            "bfb_delivery.lib.dispatch.read_circuit._get_raw_stops_list",
            return_value=mock_basic_stops_responses,
        ):
            manifest_path, circuit_sheets_dir = create_manifests_from_circuit(
                start_date=TEST_START_DATE, output_dir=output_dir
            )

        return manifest_path, circuit_sheets_dir

    @pytest.fixture(scope="class")
    def manifest_workbook(self, outputs: tuple[Path, Path]) -> Workbook:
        """Create a basic manifest workbook scoped to class for reuse."""
        workbook = load_workbook(outputs[0])
        return workbook

    @pytest.fixture(scope="class")
    def manifest_ExcelFile(self, outputs: tuple[Path, Path]) -> Iterator[pd.ExcelFile]:
        """Create a basic manifest workbook scoped to class for reuse."""
        with pd.ExcelFile(outputs[0]) as xls:
            yield xls

    @pytest.fixture(scope="class")
    def plans_list(self) -> list[dict[str, Any]]:
        """_get_raw_plans."""
        return _get_raw_plans(
            start_date=TEST_START_DATE, end_date=TEST_START_DATE, verbose=False
        )

    @pytest.fixture(scope="class")
    def plans_df(self, plans_list: list[dict[str, Any]]) -> pd.DataFrame:
        """_make_plans_df."""
        return _make_plans_df(plans_list=plans_list, all_HHs=False)

    @pytest.fixture(scope="class")
    def plan_stops_list(
        self, mock_basic_stops_responses: list, plans_df: pd.DataFrame
    ) -> list[dict[str, Any]]:
        """_get_raw_stops."""
        with patch(
            "bfb_delivery.lib.dispatch.read_circuit._get_raw_stops_list",
            return_value=mock_basic_stops_responses,
        ):
            return _get_raw_stops(
                plan_ids=plans_df[CircuitColumns.ID].tolist(), verbose=False
            )

    @pytest.fixture(scope="class")
    def transformed_routes_df(
        self, plan_stops_list: list[dict[str, Any]], plans_df: pd.DataFrame
    ) -> pd.DataFrame:
        """_transform_routes_df."""
        return _transform_routes_df(plan_stops_list=plan_stops_list, plans_df=plans_df)

    def test_date_field_matches_sheet_date(self, manifest_workbook: Workbook) -> None:
        """Test that the date field matches the sheet date."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            field_date = ws["A3"].value.split(" ")[1]
            sheet_name_date = sheet_name.split(" ")[0]
            assert field_date == sheet_name_date

    def test_df_is_same(
        self, outputs: tuple[Path, Path], manifest_ExcelFile: pd.ExcelFile
    ) -> None:
        """All the input data is in the formatted workbook."""
        for sheet_name in sorted(manifest_ExcelFile.sheet_names):
            input_df = pd.read_csv(outputs[1] / f"{sheet_name}.csv")
            output_df = pd.read_excel(manifest_ExcelFile, sheet_name=sheet_name, skiprows=8)

            # Hacky, but need to make sure formatted values haven't fundamentally changed.
            formatted_columns = [Columns.BOX_TYPE, Columns.NAME, Columns.PHONE]
            unformatted_columns = [
                col for col in FORMATTED_ROUTES_COLUMNS if col not in formatted_columns
            ]
            assert input_df[unformatted_columns].equals(output_df[unformatted_columns])

            input_df.rename(columns={Columns.PRODUCT_TYPE: Columns.BOX_TYPE}, inplace=True)
            input_box_type_df = input_df[[Columns.BOX_TYPE]]
            _format_and_validate_box_type(df=input_box_type_df)
            assert input_box_type_df.equals(output_df[[Columns.BOX_TYPE]])

            input_name_df = input_df[[Columns.NAME]]
            _format_and_validate_name(df=input_name_df)
            assert input_name_df.equals(output_df[[Columns.NAME]])

            input_phone_df = input_df[[Columns.PHONE]]
            _format_and_validate_phone(df=input_phone_df)
            assert input_phone_df.equals(output_df[[Columns.PHONE]])

    @pytest.mark.parametrize(
        "cell, expected_value",
        [
            ("A1", "DRIVER SUPPORT: 555-555-5555"),
            ("B1", None),
            ("C1", None),
            ("D1", "RECIPIENT SUPPORT: 555-555-5555 x5"),
            ("E1", None),
            ("F1", "PLEASE SHRED MANIFEST AFTER COMPLETING ROUTE."),
        ],
    )
    def test_header_row(
        self, cell: str, expected_value: str, manifest_workbook: Workbook
    ) -> None:
        """Test that the header row is correct."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            assert ws[cell].value == expected_value

    def test_header_row_end(self, manifest_workbook: Workbook) -> None:
        """Test that the header row ends at F1."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            last_non_empty_col = max(
                (cell.column for cell in ws[1] if cell.value), default=None
            )
            assert last_non_empty_col == 6

    @pytest.mark.parametrize("cell", ["A1", "B1", "C1", "D1", "E1", "F1"])
    def test_header_row_color(self, cell: str, manifest_workbook: Workbook) -> None:
        """Test the header row fill color."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            assert ws[cell].fill.start_color.rgb == f"{CellColors.HEADER}"

    def test_date_cell(self, manifest_workbook: Workbook) -> None:
        """Test that the date cell is correct."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            assert ws["A3"].value == f"Date: {MANIFEST_DATE}"

    def test_driver_cell(
        self, mock_driver_names_basic: list[str], manifest_workbook: Workbook
    ) -> None:
        """Test that the driver cell is correct."""
        drivers = [driver.upper() for driver in mock_driver_names_basic]
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            driver_name = sheet_name.replace(f"{MANIFEST_DATE} ", "")
            assert ws["A5"].value == f"Driver: {driver_name}"
            assert driver_name.upper() in drivers

    def test_agg_cells(self, manifest_workbook: Workbook, outputs: tuple[Path, Path]) -> None:
        """Test that the aggregated cells are correct."""
        for sheet_name in sorted(manifest_workbook.sheetnames):
            input_df = pd.read_csv(outputs[1] / f"{sheet_name}.csv")
            ws = manifest_workbook[sheet_name]

            input_df.rename(columns={Columns.PRODUCT_TYPE: Columns.BOX_TYPE}, inplace=True)
            agg_dict = _aggregate_route_data(
                df=input_df, extra_notes_df=get_extra_notes(file_path="")
            )

            neighborhoods = ", ".join(agg_dict["neighborhoods"])
            assert ws["A7"].value == f"Neighborhoods: {neighborhoods.upper()}"
            assert ws["E3"].value == BoxType.BASIC
            assert ws["F3"].value == agg_dict["box_counts"][BoxType.BASIC]
            assert ws["E4"].value == BoxType.GF
            assert ws["F4"].value == agg_dict["box_counts"][BoxType.GF]
            assert ws["E5"].value == BoxType.LA
            assert ws["F5"].value == agg_dict["box_counts"][BoxType.LA]
            assert ws["E6"].value == BoxType.VEGAN
            assert ws["F6"].value == agg_dict["box_counts"][BoxType.VEGAN]
            assert ws["E7"].value == "TOTAL BOX COUNT="
            assert ws["F7"].value == agg_dict["total_box_count"]
            assert ws["E8"].value == "PROTEIN COUNT="
            assert ws["F8"].value == agg_dict["protein_box_count"]

    def test_box_type_cell_colors(self, manifest_workbook: Workbook) -> None:
        """Test that the box type cells conditionally formatted with fill color."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            for cell in ws["F"]:
                if cell.row > 9:
                    assert cell.fill.start_color.rgb == f"{BOX_TYPE_COLOR_MAP[cell.value]}"
            for cell in ws["E"]:
                if cell.row > 2 and cell.row < 7:
                    assert cell.fill.start_color.rgb == f"{BOX_TYPE_COLOR_MAP[cell.value]}"

    def test_notes_column_width(self, manifest_workbook: Workbook) -> None:
        """Test that the notes column width is correct."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            assert ws["E9"].value == Columns.NOTES
            assert ws.column_dimensions["E"].width == NOTES_COLUMN_WIDTH

    @pytest.mark.parametrize(
        "cell",
        [
            # Header row.
            "A1",
            "B1",
            "C1",
            "D1",
            "E1",
            "F1",
            # Aggregated data.
            "A3",
            "A5",
            "A7",
            "E3",
            "E4",
            "E5",
            "E6",
            "E7",
            "E8",
            "F3",
            "F4",
            "F5",
            "F6",
            "F7",
            "F8",
            # Data header.
            "A9",
            "B9",
            "C9",
            "D9",
            "E9",
            "F9",
        ],
    )
    def test_bold_cells(self, cell: str, manifest_workbook: Workbook) -> None:
        """Test that the cells are bold."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            assert ws[cell].font.bold

    def test_cell_right_alignment(self, manifest_workbook: Workbook) -> None:
        """Test right-aligned cells."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            right_aligned_cells = [ws["D1"], ws["F1"]] + [
                cell for row in ws["E3:F8"] for cell in row
            ]
            for cell in right_aligned_cells:
                assert cell.alignment.horizontal == "right"

    def test_cell_left_alignment(self, manifest_workbook: Workbook) -> None:
        """Test left-aligned cells."""
        for sheet_name in manifest_workbook.sheetnames:
            ws = manifest_workbook[sheet_name]
            left_aligned_cells = [cell for row in ws["A1:A8"] for cell in row] + [
                cell for row in ws["A9:F9"] for cell in row
            ]
            for cell in left_aligned_cells:
                assert cell.alignment.horizontal == "left"

    # TODO: How to test extra notes here?

    @pytest.mark.parametrize(
        "field, bad_value, expected_error",
        [
            (
                IntermediateColumns.DRIVER_SHEET_NAME,
                None,
                pytest.raises(ValueError, match="contains null values"),
            ),
            (
                IntermediateColumns.DRIVER_SHEET_NAME,
                "OneWord",
                pytest.raises(ValueError, match="at_least_two_words"),
            ),
            (Columns.STOP_NO, np.nan, pytest.raises(ValueError, match="Could not coerce")),
            (
                Columns.STOP_NO,
                0,
                pytest.raises(ValueError, match=re.escape("greater_than_or_equal_to(1)")),
            ),
            (Columns.NAME, None, pytest.raises(ValueError, match="contains null values")),
            (Columns.ADDRESS, None, pytest.raises(ValueError, match="contains null values")),
            (
                Columns.ORDER_COUNT,
                None,
                pytest.raises(ValueError, match="contains null values"),
            ),
            (
                Columns.ORDER_COUNT,
                0,
                pytest.raises(ValueError, match=re.escape("equal_to(1)")),
            ),
            (Columns.BOX_TYPE, None, pytest.raises(ValueError, match="contains null values")),
        ],
    )
    def test_write_routes_dfs_field_checks(
        self,
        field: str,
        bad_value: Any | None,
        expected_error: AbstractContextManager,
        transformed_routes_df: pd.DataFrame,
        tmp_path_factory: pytest.TempPathFactory,
    ) -> None:
        """Raises for field violations."""
        output_dir = str(tmp_path_factory.mktemp("output"))

        bad_df = copy.deepcopy(transformed_routes_df)
        bad_df.loc[0, field] = bad_value
        with expected_error:
            _write_routes_dfs(routes_df=bad_df, output_dir=Path(output_dir))

    @pytest.mark.parametrize(
        "value, expected_error",
        [
            (
                "I am not a box type.",
                pytest.raises(ValueError, match="in_list_case_insensitive"),
            ),
            (BoxType.BASIC, nullcontext()),
            (BoxType.GF, nullcontext()),
            (BoxType.LA, nullcontext()),
            (BoxType.VEGAN, nullcontext()),
        ],
    )
    def test_write_routes_dfs_box_type_inlist(
        self,
        value: str,
        expected_error: AbstractContextManager,
        transformed_routes_df: pd.DataFrame,
        tmp_path_factory: pytest.TempPathFactory,
    ) -> None:
        """Raises error of not a real box type."""
        output_dir = str(tmp_path_factory.mktemp("output"))

        bad_df = copy.deepcopy(transformed_routes_df)
        bad_df[Columns.BOX_TYPE] = bad_df[Columns.BOX_TYPE].astype(str)
        bad_df.loc[0, Columns.BOX_TYPE] = value

        class RecastSchema(pa.DataFrameModel):
            BOX_TYPE: Series[pa.Category] = pa.Field(coerce=True, alias=Columns.BOX_TYPE)

        bad_df = RecastSchema.validate(bad_df)

        with expected_error:
            _write_routes_dfs(routes_df=bad_df, output_dir=Path(output_dir))

    # Datafrane checks are tested at a lower level to isolate them. Some checks are run first
    # and will always raise when you try to test another test.

    @pytest.mark.parametrize(
        "column", [CircuitColumns.ROUTE, IntermediateColumns.DRIVER_SHEET_NAME]
    )
    def test_write_routes_dfs_one_to_one(
        self, column: str, transformed_routes_df: pd.DataFrame
    ) -> None:
        """Raises if route:driver_sheet_name not 1:1."""
        bad_df = copy.deepcopy(transformed_routes_df)
        bad_df.loc[0, column] = bad_df.loc[len(bad_df) - 1, column]
        assert (
            one_to_one(
                df=bad_df,
                col_a=CircuitColumns.ROUTE,
                col_b=IntermediateColumns.DRIVER_SHEET_NAME,
            )
            is np.False_
        )

    @pytest.mark.parametrize(
        "group_col, at_least_one_col",
        [
            (CircuitColumns.ROUTE, IntermediateColumns.DRIVER_SHEET_NAME),
            (IntermediateColumns.DRIVER_SHEET_NAME, CircuitColumns.ROUTE),
            (IntermediateColumns.DRIVER_SHEET_NAME, Columns.STOP_NO),
        ],
    )
    def test_write_routes_dfs_at_least_one_in_group(
        self, group_col: str, at_least_one_col: str, transformed_routes_df: pd.DataFrame
    ) -> None:
        """Raises if at least one in group is not true."""
        bad_df = copy.deepcopy(transformed_routes_df)
        bad_df.loc[bad_df[group_col] == bad_df.loc[0, group_col], at_least_one_col] = None
        assert (
            at_least_one_in_group(
                df=bad_df, group_col=group_col, at_least_one_col=at_least_one_col
            )
            is False
        )

    @pytest.mark.parametrize(
        "column", [IntermediateColumns.DRIVER_SHEET_NAME, Columns.STOP_NO]
    )
    def test_write_routes_dfs_unique(
        self,
        column: str,
        transformed_routes_df: pd.DataFrame,
        tmp_path_factory: pytest.TempPathFactory,
    ) -> None:
        """Raises if driver_sheet_name:stop_no not unique."""
        output_dir = str(tmp_path_factory.mktemp("output"))

        bad_df = copy.deepcopy(transformed_routes_df)
        bad_df.loc[0, column] = bad_df.loc[len(bad_df) - 1, column]
        with pytest.raises(ValidationError, match="not unique"):
            _write_routes_dfs(routes_df=bad_df, output_dir=Path(output_dir))

    def test_write_routes_dfs_contiguous_group(
        self, transformed_routes_df: pd.DataFrame, tmp_path_factory: pytest.TempPathFactory
    ) -> None:
        """Raises if stop_no not continguous within driver_sheet_name group."""
        output_dir = str(tmp_path_factory.mktemp("output"))

        bad_df = copy.deepcopy(transformed_routes_df)
        first_group = bad_df[IntermediateColumns.DRIVER_SHEET_NAME].iloc[0]
        max_first_group = bad_df[
            bad_df[IntermediateColumns.DRIVER_SHEET_NAME] == first_group
        ][Columns.STOP_NO].max()
        bad_df.loc[
            (bad_df[IntermediateColumns.DRIVER_SHEET_NAME] == first_group)
            & (bad_df[Columns.STOP_NO] == max_first_group),
            Columns.STOP_NO,
        ] = (max_first_group + 1)

        with pytest.raises(ValidationError, match="contiguous_group"):
            _write_routes_dfs(routes_df=bad_df, output_dir=Path(output_dir))

    def test_write_routes_dfs_increasing_by(
        self, transformed_routes_df: pd.DataFrame, tmp_path_factory: pytest.TempPathFactory
    ) -> None:
        """Raises if not sorted by stop_no within driver_sheet_name."""
        output_dir = str(tmp_path_factory.mktemp("output"))

        bad_df = copy.deepcopy(transformed_routes_df)
        first_group = bad_df[IntermediateColumns.DRIVER_SHEET_NAME].iloc[0]
        first_group_vals = bad_df[
            bad_df[IntermediateColumns.DRIVER_SHEET_NAME] == first_group
        ][Columns.STOP_NO].to_list()
        bad_df.loc[
            bad_df[IntermediateColumns.DRIVER_SHEET_NAME] == first_group, Columns.STOP_NO
        ] = list(reversed(first_group_vals))

        with pytest.raises(ValidationError, match="increasing_by"):
            _write_routes_dfs(routes_df=bad_df, output_dir=Path(output_dir))

    def test_write_routes_dfs_many_to_one(
        self, transformed_routes_df: pd.DataFrame, tmp_path_factory: pytest.TempPathFactory
    ) -> None:
        """Raises if driver_sheet_name:stop_id not m:1."""
        output_dir = str(tmp_path_factory.mktemp("output"))

        bad_df = copy.deepcopy(transformed_routes_df)
        bad_df.loc[0, CircuitColumns.ID] = bad_df.loc[len(bad_df) - 1, CircuitColumns.ID]
        with pytest.raises(ValidationError, match="many_to_one"):
            _write_routes_dfs(routes_df=bad_df, output_dir=Path(output_dir))

    @pytest.mark.parametrize(
        "field, item, check_name",
        [
            (CircuitColumns.ADDRESS, CircuitColumns.ADDRESS_LINE_1, "address1_in_address"),
            (CircuitColumns.ADDRESS, CircuitColumns.ADDRESS_LINE_2, "address2_in_address"),
            (CircuitColumns.ADDRESS, CircuitColumns.PLACE_ID, "item_in_field_dict"),
            (CircuitColumns.ROUTE, CircuitColumns.ID, "item_in_field_dict"),
            (CircuitColumns.RECIPIENT, CircuitColumns.NAME, "item_in_field_dict"),
            (CircuitColumns.ORDER_INFO, CircuitColumns.PRODUCTS, "item_in_field_dict"),
        ],
    )
    def test_transform_routes_df_item_in_field_dict(
        self,
        field: str,
        item: str,
        check_name: str,
        plans_df: pd.DataFrame,
        plan_stops_list: list[dict[str, Any]],
    ) -> None:
        """Raises if item not in address column dict."""
        bad_plan_stops_list = copy.deepcopy(plan_stops_list)
        bad_plan_stops_list[0][field].pop(item)
        with pytest.raises(ValidationError, match=check_name):
            _transform_routes_df(plan_stops_list=bad_plan_stops_list, plans_df=plans_df)

    @pytest.mark.parametrize(
        "field, bad_value, expected_error",
        [
            (
                CircuitColumns.PLAN,
                None,
                pytest.raises(SchemaError, match="contains null values"),
            ),
            (
                CircuitColumns.ROUTE,
                None,
                pytest.raises(SchemaError, match="contains null values"),
            ),
            (
                CircuitColumns.ID,
                None,
                pytest.raises(SchemaError, match="contains null values"),
            ),
            (
                CircuitColumns.STOP_POSITION,
                None,
                pytest.raises(SchemaError, match="Could not coerce"),
            ),
            (
                CircuitColumns.STOP_POSITION,
                -1,
                pytest.raises(SchemaError, match=re.escape("greater_than_or_equal_to(0)")),
            ),
            (
                CircuitColumns.RECIPIENT,
                None,
                pytest.raises(SchemaError, match="contains null values"),
            ),
            (
                CircuitColumns.ADDRESS,
                None,
                pytest.raises(SchemaError, match="contains null values"),
            ),
            (
                CircuitColumns.PACKAGE_COUNT,
                0,
                pytest.raises(SchemaError, match=re.escape("equal_to(1)")),
            ),
        ],
    )
    def test_transform_routes_df_field_checks(
        self,
        field: str,
        bad_value: Any | None,
        expected_error: AbstractContextManager,
        plan_stops_list: list[dict[str, Any]],
    ) -> None:
        """Raises for field violations."""
        bad_df = pd.DataFrame(plan_stops_list)
        bad_df.loc[0, field] = bad_value
        with expected_error:
            _ = CircuitRoutesTransformInFromDict.validate(bad_df)

    def test_transform_routes_df_stop_id_unique(
        self, plan_stops_list: list[dict[str, Any]]
    ) -> None:
        """Raises if id not unique."""
        bad_df = pd.DataFrame(plan_stops_list)
        bad_df.loc[0, CircuitColumns.ID] = bad_df.loc[1, CircuitColumns.ID]
        with pytest.raises(SchemaError, match="contains duplicate value"):
            _ = CircuitRoutesTransformInFromDict.validate(bad_df)

    @pytest.mark.parametrize(
        "id, error_match",
        [("asdg/stops/asdg", "str_startswith"), ("plans/afasdf/sdfsa", "str_contains")],
    )
    def test_transform_routes_df_stop_id_regex(
        self, id: str, error_match: str, plan_stops_list: list[dict[str, Any]]
    ) -> None:
        """Raises if id doesn't fit pattern."""
        bad_df = pd.DataFrame(plan_stops_list)
        bad_df.loc[0, CircuitColumns.ID] = id
        with pytest.raises(SchemaError, match=error_match):
            _ = CircuitRoutesTransformInFromDict.validate(bad_df)


@pytest.mark.parametrize(
    "mock_responses, expected_result",
    [
        (
            [
                {
                    "json.return_value": {"data": [1, 2, 3], "nextPageToken": None},
                    "status_code": 200,
                }
            ],
            [{"data": [1, 2, 3], "nextPageToken": None}],
        ),
        (
            [
                {
                    "json.return_value": {"data": [1], "nextPageToken": "abc"},
                    "status_code": 200,
                },
                {
                    "json.return_value": {"data": [2], "nextPageToken": None},
                    "status_code": 200,
                },
            ],
            [{"data": [1], "nextPageToken": "abc"}, {"data": [2], "nextPageToken": None}],
        ),
        (
            [
                {"json.return_value": {}, "status_code": 429},
                {
                    "json.return_value": {"data": [3], "nextPageToken": None},
                    "status_code": 200,
                },
            ],
            [{"data": [3], "nextPageToken": None}],
        ),
        ([{"json.return_value": {}, "status_code": 400}], requests.exceptions.HTTPError),
        ([{"json.return_value": {}, "status_code": 401}], requests.exceptions.HTTPError),
        ([{"json.return_value": {}, "status_code": 403}], requests.exceptions.HTTPError),
        ([{"json.return_value": {}, "status_code": 404}], requests.exceptions.HTTPError),
        ([{"json.return_value": {}, "status_code": 500}], requests.exceptions.HTTPError),
    ],
)
def test_get_responses(mock_responses: list[dict[str, Any]], expected_result: Any) -> None:
    """Test _get_responses function."""
    url = "https://fakeapi.com/data"

    with patch("bfb_delivery.lib.dispatch.read_circuit.requests.get") as mock_get:
        mock_get.side_effect = [Mock(**resp) for resp in mock_responses]

        if isinstance(expected_result, type) and issubclass(expected_result, Exception):
            with pytest.raises(expected_result):
                _ = _get_responses(url)
        else:
            result = _get_responses(url)
            assert result == expected_result

        assert mock_get.call_count == len(mock_responses)
