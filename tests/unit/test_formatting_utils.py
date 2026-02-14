"""Unit tests for the utility functions."""

import logging
import re
from contextlib import AbstractContextManager, nullcontext
from pathlib import Path
from typing import Final, cast
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from typeguard import typechecked

from bfb_delivery.lib.constants import LINE_HEIGHT
from bfb_delivery.lib.formatting.utils import (
    get_extra_notes,
    get_phone_number,
    map_columns,
    set_row_height_of_wrapped_cell,
)


@pytest.mark.parametrize(
    "key, config_path, expected, expected_warning",
    [
        ("driver_support", "config.ini", "555-555-5555", ""),
        ("recipient_support", "config.ini", "555-555-5555 x5", ""),
        (
            "bad_key",
            "config.ini",
            (
                "NO PHONE NUMBER. See warning in logs for instructions on setting up your "
                "config file."
            ),
            "bad_key not found in config file: ",
        ),
        ("driver_support", None, "555-555-5555", ""),
        ("recipient_support", None, "555-555-5555 x5", ""),
        (
            "bad_key",
            None,
            (
                "NO PHONE NUMBER. See warning in logs for instructions on setting up your "
                "config file."
            ),
            "bad_key not found in config file: ",
        ),
        (
            "driver_support",
            "bad_config.ini",
            (
                "NO PHONE NUMBER. See warning in logs for instructions on setting up your "
                "config file."
            ),
            "Config file not found: ",
        ),
        (
            "recipient_support",
            "bad_config.ini",
            (
                "NO PHONE NUMBER. See warning in logs for instructions on setting up your "
                "config file."
            ),
            "Config file not found: ",
        ),
        (
            "bad_key",
            "bad_config.ini",
            (
                "NO PHONE NUMBER. See warning in logs for instructions on setting up your "
                "config file."
            ),
            "Config file not found: ",
        ),
    ],
)
@typechecked
def test_get_phone_number(
    key: str,
    config_path: None | str,
    expected: str,
    expected_warning: str,
    caplog: pytest.LogCaptureFixture,
) -> None:
    """Correct phone number returned, and warns on no config file or missing key."""
    kwargs = {"key": key}
    if config_path is not None:
        kwargs["config_path"] = config_path
    if expected_warning:
        with caplog.at_level(logging.WARNING):
            returned_phone_number = get_phone_number(**kwargs)
            assert expected_warning in caplog.text
    else:
        returned_phone_number = get_phone_number(**kwargs)

    assert returned_phone_number == expected


@pytest.mark.parametrize("invert_map", [False, True])
@typechecked
def test_map_columns(invert_map: bool) -> None:
    """Test that the columns are mapped correctly."""
    df = pd.DataFrame(
        {
            "Box Type": ["BASIC", "GF", "LA"],
            "Box Count": [1, 2, 3],
            "Driver": ["John", "Jane", "Jim"],
        }
    )
    column_name_map = {"Box Type": "Product Type"}
    map_columns(df, column_name_map, invert_map=invert_map)
    if not invert_map:
        assert df.columns.to_list() == ["Product Type", "Box Count", "Driver"]
    else:
        assert df.columns.to_list() == ["Box Type", "Box Count", "Driver"]


@pytest.mark.parametrize(
    "extra_notes_df, error_context",
    [
        (
            pd.DataFrame(
                columns=["tag", "note"],
                data=[("tag1", "note1"), ("tag2", "note2"), ("tag3", "note3")],
            ),
            nullcontext(),
        ),
        (
            pd.DataFrame(
                columns=["tag", "note"],
                data=[("tag1", "note1"), ("tag2", "note2"), ("tag1", "note3")],
            ),
            pytest.raises(
                ValueError, match=re.escape(("Extra notes has duplicated tags: ['tag1']"))
            ),
        ),
        (
            pd.DataFrame(
                columns=["tag", "note"],
                data=[("tag1*", "note1"), ("tag2", "note2"), ("tag1 *", "note3")],
            ),
            pytest.raises(
                ValueError, match=re.escape(("Extra notes has duplicated tags: ['tag1']"))
            ),
        ),
    ],
)
@pytest.mark.parametrize("file_name", ["extra_notes.csv", ""])
@typechecked
def test_get_extra_notes(
    extra_notes_df: pd.DataFrame,
    error_context: AbstractContextManager,
    file_name: str,
    tmp_path: Path,
) -> None:
    """Test that the extra notes are read correctly."""
    file_path = str(tmp_path / file_name) if file_name else file_name

    mock_extra_notes_context: AbstractContextManager

    if file_name:
        mock_extra_notes_context = nullcontext()
        extra_notes_df.to_csv(file_path, index=False)
    else:

        class TestExtraNotes:
            df: Final[pd.DataFrame] = extra_notes_df

        mock_extra_notes_context = patch(
            "bfb_delivery.lib.formatting.utils.ExtraNotes", new=TestExtraNotes
        )

    with error_context, mock_extra_notes_context:
        returned_extra_notes_df = get_extra_notes(file_path=file_path)
        assert returned_extra_notes_df.equals(extra_notes_df)


@pytest.mark.parametrize("bold", [False, True])
@pytest.mark.parametrize(
    "cell_value, expected_height_not_bold, expected_height_bold",
    [
        ("Short text", LINE_HEIGHT, LINE_HEIGHT),
        (
            "This is a test string that's over one hundred characters long to "
            "test bold text wrapping .............",
            LINE_HEIGHT,
            LINE_HEIGHT * 2,
        ),
        (
            "This is an extremely long text that contains a lot of information and will "
            "definitely need to wrap across multiple lines when displayed in a cell",
            LINE_HEIGHT * 2,
            LINE_HEIGHT * 2,
        ),
    ],
)
@typechecked
def test_set_row_height_of_wrapped_cell(
    cell_value: str, expected_height_not_bold: float, expected_height_bold: float, bold: bool
) -> None:
    """Test row height calculation for cells with wrapped text."""
    wb = Workbook()
    ws = cast(Worksheet, wb.active)

    for col in range(1, 7):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    cell = ws.cell(row=1, column=1, value=cell_value)
    if bold:
        cell.font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    set_row_height_of_wrapped_cell(cell=cell)

    expected_height = expected_height_bold if bold else expected_height_not_bold
    height = ws.row_dimensions[1].height
    assert height == expected_height
    assert height % LINE_HEIGHT == 0
