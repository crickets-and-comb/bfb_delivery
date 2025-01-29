"""Unit tests for sheet_shaping.py."""

# TODO: Make fixtures class-scoped again, at least for the basic calls, here and elsewhere.
import glob
import re
import subprocess
from collections.abc import Iterator
from contextlib import AbstractContextManager, nullcontext
from datetime import datetime
from enum import StrEnum
from pathlib import Path
from typing import Final
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from bfb_delivery import (
    combine_route_tables,
    create_manifests,
    format_combined_routes,
    split_chunked_route,
)
from bfb_delivery.lib.constants import (
    BOX_TYPE_COLOR_MAP,
    COMBINED_ROUTES_COLUMNS,
    FILE_DATE_FORMAT,
    FORMATTED_ROUTES_COLUMNS,
    MANIFEST_DATE_FORMAT,
    NOTES_COLUMN_WIDTH,
    SPLIT_ROUTE_COLUMNS,
    BoxType,
    CellColors,
    Columns,
    Defaults,
)
from bfb_delivery.lib.formatting.data_cleaning import (
    _format_and_validate_box_type,
    _format_and_validate_name,
    _format_and_validate_neighborhood,
    _format_and_validate_phone,
)
from bfb_delivery.lib.formatting.sheet_shaping import (
    _aggregate_route_data,
    _get_driver_sets,
    _group_numbered_drivers,
)
from bfb_delivery.lib.formatting.utils import get_extra_notes

BOX_TYPES: Final[list[str]] = ["Basic", "GF", "Vegan", "LA"]
DRIVERS: Final[list[str]] = [
    "Driver A",
    "Driver B",
    "Driver C",
    "Driver D #1",
    "Driver D #2",
    "Driver E",
    "Driver F",
]
MANIFEST_DATE: Final[str] = "1.1"
N_BOOKS_MATRIX: Final[list[int]] = [1, 3, 4]
NEIGHBORHOODS: Final[list[str]] = ["York", "Puget", "Samish", "Sehome", "South Hill"]


@pytest.fixture()
def mock_chunked_sheet_raw(tmp_path: Path) -> Path:
    """Save mock chunked route sheet and get path."""
    fp: Path = tmp_path / "mock_chunked_sheet_raw.xlsx"
    raw_chunked_sheet = pd.DataFrame(
        columns=SPLIT_ROUTE_COLUMNS + [Columns.DRIVER, Columns.BOX_COUNT, Columns.STOP_NO],
        data=[
            (
                "Recipient One",
                "123 Main St",
                "555-555-1234",
                "Recipient1@email.com",
                "Notes for Recipient One.",
                "1",
                "Basic",
                "York",
                "Driver A",
                2,
                1,
            ),
            (
                "Recipient Two",
                "456 Elm St",
                "555-555-5678",
                "Recipient2@email.com",
                "Notes for Recipient Two.",
                "1",
                "GF",
                "Puget",
                "Driver A",
                None,
                2,
            ),
            (
                "Recipient Three",
                "789 Oak St",
                "555-555-9101",
                "Recipient3@email.com",
                "Notes for Recipient Three.",
                "1",
                "Vegan",
                "Puget",
                "Driver B",
                2,
                3,
            ),
            (
                "Recipient Four",
                "1011 Pine St",
                "555-555-1121",
                "Recipient4@email.com",
                "Notes for Recipient Four.",
                "1",
                "LA",
                "Puget",
                "Driver B",
                None,
                4,
            ),
            (
                "Recipient Five",
                "1314 Cedar St",
                "555-555-3141",
                "Recipient5@email.com",
                "Notes for Recipient Five.",
                "1",
                "Basic",
                "Samish",
                "Driver C",
                1,
                5,
            ),
            (
                "Recipient Six",
                "1516 Fir St",
                "555-555-5161",
                "Recipient6@email.com",
                "Notes for Recipient Six.",
                "1",
                "GF",
                "Sehome",
                "Driver D #1",
                1,
                6,
            ),
            (
                "Recipient Seven",
                "1718 Spruce St",
                "555-555-7181",
                "Recipient7@email.com",
                "Notes for Recipient Seven.",
                "1",
                "Vegan",
                "Samish",
                "Driver D #2",
                2,
                7,
            ),
            (
                "Recipient Eight",
                "1920 Maple St",
                "555-555-9202",
                "Recipient8@email.com",
                "Notes for Recipient Eight.",
                "1",
                "LA",
                "South Hill",
                "Driver D #2",
                None,
                8,
            ),
            (
                "Recipient Nine",
                "2122 Cedar St",
                "555-555-2223",
                "Recipient9@email.com",
                "Notes for Recipient Nine.",
                "1",
                "Basic",
                "South Hill",
                "Driver E",
                2,
                9,
            ),
            (
                "Recipient Ten",
                "2122 Cedar St",
                "555-555-2223",
                "Recipient10@email.com",
                "Notes for Recipient Ten.",
                "1",
                "LA",
                "South Hill",
                "Driver E",
                None,
                10,
            ),
            (
                "Recipient Eleven",
                "2346 Ash St",
                "555-555-2345",
                "Recipient11@email.com",
                "Notes for Recipient Eleven.",
                "1",
                "Basic",
                "Eldridge",
                "Driver F",
                2,
                11,
            ),
            (
                "Recipient Twelve",
                "2122 Cedar St",
                "555-555-2223",
                "Recipient12@email.com",
                "Notes for Recipient Twelve.",
                "1",
                "Basic",
                "Eldridge",
                "Driver F",
                None,
                12,
            ),
        ],
    ).rename(columns={Columns.PRODUCT_TYPE: Columns.BOX_TYPE})
    raw_chunked_sheet.to_excel(fp, index=False)

    return fp


@pytest.fixture()
def mock_route_tables(tmp_path: Path, mock_chunked_sheet_raw: Path) -> Path:
    """Mock the driver route tables returned by Circuit."""
    output_dir = tmp_path / "mock_route_tables"
    output_dir.mkdir()

    output_cols = [Columns.STOP_NO] + SPLIT_ROUTE_COLUMNS
    chunked_df = pd.read_excel(mock_chunked_sheet_raw)
    chunked_df.rename(columns={Columns.BOX_TYPE: Columns.PRODUCT_TYPE}, inplace=True)
    for driver in chunked_df[Columns.DRIVER].unique():
        output_path = output_dir / f"{MANIFEST_DATE} {driver}.csv"
        driver_df = chunked_df[chunked_df[Columns.DRIVER] == driver]
        driver_df[Columns.STOP_NO] = [i + 1 for i in range(len(driver_df))]
        driver_df[output_cols].to_csv(output_dir / output_path, index=False)

    return output_dir


@pytest.fixture(scope="class")
def mock_chunked_sheet_raw_class_scoped(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """Save mock chunked route sheet and get path."""
    tmp_output = tmp_path_factory.mktemp(
        "tmp_mock_chunked_sheet_raw_class_scoped", numbered=True
    )
    fp: Path = tmp_output / "mock_chunked_sheet_raw.xlsx"
    raw_chunked_sheet = pd.DataFrame(
        columns=SPLIT_ROUTE_COLUMNS + [Columns.DRIVER, Columns.BOX_COUNT, Columns.STOP_NO],
        data=[
            (
                "Recipient One",
                "123 Main St",
                "555-555-1234",
                "Recipient1@email.com",
                "Notes for Recipient One.",
                "1",
                "Basic",
                "York",
                "Driver A",
                2,
                1,
            ),
            (
                "Recipient Two",
                "456 Elm St",
                "555-555-5678",
                "Recipient2@email.com",
                "Notes for Recipient Two.",
                "1",
                "GF",
                "Puget",
                "Driver A",
                None,
                2,
            ),
            (
                "Recipient Three",
                "789 Oak St",
                "555-555-9101",
                "Recipient3@email.com",
                "Notes for Recipient Three.",
                "1",
                "Vegan",
                "Puget",
                "Driver B",
                2,
                3,
            ),
            (
                "Recipient Four",
                "1011 Pine St",
                "555-555-1121",
                "Recipient4@email.com",
                "Notes for Recipient Four.",
                "1",
                "LA",
                "Puget",
                "Driver B",
                None,
                4,
            ),
            (
                "Recipient Five",
                "1314 Cedar St",
                "555-555-3141",
                "Recipient5@email.com",
                "Notes for Recipient Five.",
                "1",
                "Basic",
                "Samish",
                "Driver C",
                1,
                5,
            ),
            (
                "Recipient Six",
                "1516 Fir St",
                "555-555-5161",
                "Recipient6@email.com",
                "Notes for Recipient Six.",
                "1",
                "GF",
                "Sehome",
                "Driver D #1",
                1,
                6,
            ),
            (
                "Recipient Seven",
                "1718 Spruce St",
                "555-555-7181",
                "Recipient7@email.com",
                "Notes for Recipient Seven.",
                "1",
                "Vegan",
                "Samish",
                "Driver D #2",
                2,
                7,
            ),
            (
                "Recipient Eight",
                "1920 Maple St",
                "555-555-9202",
                "Recipient8@email.com",
                "Notes for Recipient Eight.",
                "1",
                "LA",
                "South Hill",
                "Driver D #2",
                None,
                8,
            ),
            (
                "Recipient Nine",
                "2122 Cedar St",
                "555-555-2223",
                "Recipient9@email.com",
                "Notes for Recipient Nine.",
                "1",
                "Basic",
                "South Hill",
                "Driver E",
                2,
                9,
            ),
            (
                "Recipient Ten",
                "2122 Cedar St",
                "555-555-2223",
                "Recipient10@email.com",
                "Notes for Recipient Ten.",
                "1",
                "LA",
                "South Hill",
                "Driver E",
                None,
                10,
            ),
            (
                "Recipient Eleven",
                "2346 Ash St",
                "555-555-2345",
                "Recipient11@email.com",
                "Notes for Recipient Eleven.",
                "1",
                "Basic",
                "Eldridge",
                "Driver F",
                2,
                11,
            ),
            (
                "Recipient Twelve",
                "2122 Cedar St",
                "555-555-2223",
                "Recipient12@email.com",
                "Notes for Recipient Twelve.",
                "1",
                "Basic",
                "Eldridge",
                "Driver F",
                None,
                12,
            ),
        ],
    ).rename(columns={Columns.PRODUCT_TYPE: Columns.BOX_TYPE})
    raw_chunked_sheet.to_excel(fp, index=False)

    return fp


@pytest.fixture(scope="class")
def mock_route_tables_class_scoped(
    tmp_path_factory: pytest.TempPathFactory, mock_chunked_sheet_raw_class_scoped: Path
) -> Path:
    """Mock the driver route tables returned by Circuit."""
    tmp_output = tmp_path_factory.mktemp("tmp_mock_route_tables_class_scoped", numbered=True)
    output_dir = tmp_output / "mock_route_tables"
    output_dir.mkdir()

    output_cols = [Columns.STOP_NO] + SPLIT_ROUTE_COLUMNS
    chunked_df = pd.read_excel(mock_chunked_sheet_raw_class_scoped)
    chunked_df.rename(columns={Columns.BOX_TYPE: Columns.PRODUCT_TYPE}, inplace=True)
    for driver in chunked_df[Columns.DRIVER].unique():
        output_path = output_dir / f"{MANIFEST_DATE} {driver}.csv"
        driver_df = chunked_df[chunked_df[Columns.DRIVER] == driver]
        driver_df[Columns.STOP_NO] = [i + 1 for i in range(len(driver_df))]
        driver_df[output_cols].to_csv(output_dir / output_path, index=False)

    return output_dir


@pytest.fixture()
def mock_extra_notes_df() -> pd.DataFrame:
    """Mock the extra notes DataFrame."""
    extra_notes_df = pd.DataFrame(
        columns=["tag", "note"],
        data=[
            (
                "Test extra notes tag 1 *",
                (
                    "Test extra notes note 1. "
                    "This is a dummy note. It is really long and should be so that we can "
                    "test out column width and word wrapping. It should be long enough to "
                    "wrap around to the next line. And, it should be long enough to wrap "
                    "around to the next line. And, it should be long enough to wrap around "
                    "to the next line. Hopefully, this is long enough. Also, hopefully, this "
                    "is long enough. Further, hopefully, this is long enough. Additionally, "
                    "it will help test out word wrapping merged cells."
                ),
            ),
            ("Test extra notes tag 2 *", "Test extra notes note 2"),
            ("Test extra notes tag 3 *", "Test extra notes note 3"),
        ],
    )
    return extra_notes_df


@pytest.mark.usefixtures("mock_is_valid_number")
class TestSplitChunkedRoute:
    """split_chunked_route splits route spreadsheet into n workbooks with sheets by driver."""

    @pytest.mark.parametrize("output_dir_type", [Path, str])
    @pytest.mark.parametrize("output_dir", ["", "dummy_output"])
    @pytest.mark.parametrize("n_books", [1, 4])
    def test_set_output_dir(
        self,
        output_dir_type: type[Path | str],
        output_dir: Path | str,
        n_books: int,
        mock_chunked_sheet_raw_class_scoped: Path,
        tmp_path: Path,
    ) -> None:
        """Test that the output directory can be set."""
        output_dir = output_dir_type(tmp_path / output_dir)
        output_paths = split_chunked_route(
            input_path=mock_chunked_sheet_raw_class_scoped,
            output_dir=output_dir,
            n_books=n_books,
        )
        assert all(str(output_path.parent) == str(output_dir) for output_path in output_paths)

    @pytest.mark.parametrize("output_filename", ["", "dummy_output_filename.xlsx"])
    @pytest.mark.parametrize("n_books", [1, 4])
    def test_set_output_filename(
        self,
        output_filename: str,
        mock_chunked_sheet_raw_class_scoped: Path,
        n_books: int,
        tmp_path: Path,
    ) -> None:
        """Test that the output filename can be set."""
        output_paths = split_chunked_route(
            output_dir=tmp_path,
            input_path=mock_chunked_sheet_raw_class_scoped,
            output_filename=output_filename,
            n_books=n_books,
        )
        for i, output_path in enumerate(output_paths):
            expected_filename = output_filename.split(".")[0]
            expected_filename = (
                f"{expected_filename}_{i + 1}.xlsx"
                if output_filename
                else (
                    f"split_workbook_{datetime.now().strftime(FILE_DATE_FORMAT)}"
                    f"_{i + 1}.xlsx"
                )
            )
            assert output_path.name == expected_filename

    @pytest.mark.parametrize("n_books_passed", N_BOOKS_MATRIX + [None])
    def test_n_books_count(
        self,
        n_books_passed: int | None,
        mock_chunked_sheet_raw_class_scoped: Path,
        tmp_path: Path,
    ) -> None:
        """Test that the number of workbooks is equal to n_books."""
        if n_books_passed is None:
            output_paths = split_chunked_route(
                output_dir=tmp_path, input_path=mock_chunked_sheet_raw_class_scoped
            )
            n_books = Defaults.SPLIT_CHUNKED_ROUTE["n_books"]
        else:
            n_books = n_books_passed
            output_paths = split_chunked_route(
                output_dir=tmp_path,
                input_path=mock_chunked_sheet_raw_class_scoped,
                n_books=n_books,
            )

        assert len(output_paths) == n_books

    @pytest.mark.parametrize("n_books", N_BOOKS_MATRIX)
    def test_recipients_unique(
        self, n_books: int, mock_chunked_sheet_raw_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the recipients don't overlap between the split workbooks.

        By name, address, phone, and email.
        """
        output_paths = split_chunked_route(
            output_dir=tmp_path,
            input_path=mock_chunked_sheet_raw_class_scoped,
            n_books=n_books,
        )

        recipient_sets = []
        for output_path in output_paths:
            driver_sheets = _get_driver_sheets(output_paths=[output_path])
            recipient_sets.append(
                pd.concat(driver_sheets, ignore_index=True)[
                    [Columns.NAME, Columns.ADDRESS, Columns.PHONE, Columns.EMAIL]
                ]
            )
        recipients_df = pd.concat(recipient_sets, ignore_index=True)
        assert recipients_df.duplicated().sum() == 0

    @pytest.mark.parametrize("n_books", N_BOOKS_MATRIX)
    def test_unique_drivers_across_books(
        self, n_books: int, mock_chunked_sheet_raw_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the drivers don't overlap between the split workbooks."""
        output_paths = split_chunked_route(
            output_dir=tmp_path,
            input_path=mock_chunked_sheet_raw_class_scoped,
            n_books=n_books,
        )

        driver_sets = []
        for output_path in output_paths:
            drivers = [
                " ".join(str(sheet_name).split(" ")[1:])
                for sheet_name in pd.ExcelFile(output_path).sheet_names
            ]
            driver_sets.append(drivers)
        for i, driver_set in enumerate(driver_sets):
            driver_sets_sans_i = driver_sets[:i] + driver_sets[i + 1 :]  # noqa: E203
            driver_sets_sans_i = [
                driver for sublist in driver_sets_sans_i for driver in sublist
            ]
            assert len(set(driver_set).intersection(set(driver_sets_sans_i))) == 0

    @pytest.mark.parametrize("n_books", N_BOOKS_MATRIX)
    def test_numbered_drivers_grouped(
        self, n_books: int, mock_chunked_sheet_raw_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the numbered drivers are in the same workbook together."""
        output_paths = split_chunked_route(
            output_dir=tmp_path,
            input_path=mock_chunked_sheet_raw_class_scoped,
            n_books=n_books,
        )
        driver_d_sheets_found = False
        for output_path in output_paths:
            driver_d_sheets = [
                sheet_name
                for sheet_name in pd.ExcelFile(output_path).sheet_names
                if "Driver D" in str(sheet_name)
            ]
            assert len(driver_d_sheets) == 2 or len(driver_d_sheets) == 0

            if driver_d_sheets:
                driver_d_sheets_found = True

        assert driver_d_sheets_found

    @pytest.mark.parametrize("n_books", [1, 2, 3])
    @pytest.mark.parametrize(
        "test_book_one_drivers, exclude_drivers",
        [
            (["Driver A", "Driver B"], []),
            (["Driver F"], []),
            ([], []),
            (
                ["Driver A", "Driver B", "not a driver in the data"],
                ["not a driver in the data"],
            ),
        ],
    )
    @pytest.mark.parametrize("book_one_drivers_file", ["", "dummy_book_one_drivers.csv"])
    def test_book_one_drivers(
        self,
        n_books: int,
        test_book_one_drivers: list[str],
        exclude_drivers: list[str],
        book_one_drivers_file: str,
        mock_chunked_sheet_raw_class_scoped: Path,
        tmp_path: Path,
    ) -> None:
        """Test that book-one drivers are in book one."""
        if test_book_one_drivers:
            TestBookOneDrivers = StrEnum(
                "TestBookOneDrivers", {driver: driver for driver in test_book_one_drivers}
            )
        else:

            class TestBookOneDrivers(StrEnum):
                pass

        mock_constant_context = (
            patch("bfb_delivery.lib.formatting.utils.BookOneDrivers", new=TestBookOneDrivers)
            if not book_one_drivers_file
            else nullcontext()
        )
        book_one_drivers_file_path = (
            str(tmp_path / book_one_drivers_file)
            if book_one_drivers_file
            else book_one_drivers_file
        )
        if book_one_drivers_file_path:
            with open(book_one_drivers_file_path, "w") as f:
                f.write(f"{Columns.DRIVER}\n")
                for driver in test_book_one_drivers:
                    f.write(f"{driver}\n")

        mock_chunked_sheet_raw_df = pd.read_excel(mock_chunked_sheet_raw_class_scoped)
        alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        mock_chunked_sheet_raw_df[Columns.DRIVER] = [
            f"Driver {alphabet[i]}" for i in range(len(mock_chunked_sheet_raw_df))
        ]
        new_mock_chunked_sheet_raw_path = tmp_path / "new_mock_chunked_sheet_raw.xlsx"
        mock_chunked_sheet_raw_df.to_excel(new_mock_chunked_sheet_raw_path, index=False)

        output_dir = tmp_path / "output"
        output_dir.mkdir()
        with mock_constant_context:
            output_paths = split_chunked_route(
                output_dir=output_dir,
                input_path=new_mock_chunked_sheet_raw_path,
                n_books=n_books,
                book_one_drivers_file=book_one_drivers_file_path,
            )
        book_one = pd.ExcelFile(output_paths[0])
        book_one_drivers = [
            " ".join(str(sheet_name).split(" ")[1:]) for sheet_name in book_one.sheet_names
        ]

        missing_drivers = (
            set(test_book_one_drivers) - set(book_one_drivers) - set(exclude_drivers)
        )
        assert len(missing_drivers) == 0

        misincluded_drivers = set(book_one_drivers).intersection(set(exclude_drivers))
        assert len(misincluded_drivers) == 0

    @pytest.mark.parametrize("n_books", [1, 2, 3])
    def test_complete_contents(
        self, n_books: int, mock_chunked_sheet_raw_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the input data is all covered in the split workbooks."""
        output_paths = split_chunked_route(
            output_dir=tmp_path,
            input_path=mock_chunked_sheet_raw_class_scoped,
            n_books=n_books,
        )

        full_data = pd.read_excel(mock_chunked_sheet_raw_class_scoped)

        driver_sheets = _get_driver_sheets(output_paths=output_paths)
        split_data = pd.concat(driver_sheets, ignore_index=True)

        split_data.rename(columns={Columns.PRODUCT_TYPE: Columns.BOX_TYPE}, inplace=True)
        cols = split_data.columns.to_list()
        full_data = full_data[cols].sort_values(by=cols).reset_index(drop=True)
        split_data = split_data.sort_values(by=cols).reset_index(drop=True)

        # Hacky, but need to make sure formatted values haven't fundamentally changed.
        cols_without_formatting = [
            col
            for col in cols
            if col
            not in [Columns.PHONE, Columns.NAME, Columns.BOX_TYPE, Columns.NEIGHBORHOOD]
        ]
        pd.testing.assert_frame_equal(
            full_data[cols_without_formatting], split_data[cols_without_formatting]
        )

        phone_df = full_data[[Columns.PHONE]].copy()
        _format_and_validate_phone(df=phone_df)
        assert phone_df[Columns.PHONE].equals(split_data[Columns.PHONE])

        name_df = full_data[[Columns.NAME]].copy()
        _format_and_validate_name(df=name_df)
        assert name_df[Columns.NAME].equals(split_data[Columns.NAME])

        box_type_df = full_data[[Columns.BOX_TYPE]].copy()
        _format_and_validate_box_type(df=box_type_df)
        assert box_type_df[Columns.BOX_TYPE].equals(split_data[Columns.BOX_TYPE])

        neighborhood_df = full_data[[Columns.NEIGHBORHOOD]].copy()
        _format_and_validate_neighborhood(df=neighborhood_df)
        assert neighborhood_df[Columns.NEIGHBORHOOD].equals(split_data[Columns.NEIGHBORHOOD])

    @pytest.mark.parametrize("n_books", [0, -1])
    def test_invalid_n_books(
        self, n_books: int, mock_chunked_sheet_raw_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that an invalid n_books raises a ValueError."""
        with pytest.raises(ValueError, match="n_books must be greater than 0."):
            _ = split_chunked_route(
                output_dir=tmp_path,
                input_path=mock_chunked_sheet_raw_class_scoped,
                n_books=n_books,
            )

    def test_invalid_n_books_driver_count(
        self, mock_chunked_sheet_raw_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that n_books greater than the number of drivers raises a ValueError."""
        raw_sheet = pd.read_excel(mock_chunked_sheet_raw_class_scoped)
        driver_count = len(raw_sheet[Columns.DRIVER].unique())
        n_books = driver_count + 1
        with pytest.raises(
            ValueError,
            match=(
                f"n_books must be less than or equal to the number of drivers: "
                f"driver_count: ({driver_count}), n_books: {n_books}."
            ),
        ):
            _ = split_chunked_route(
                output_dir=tmp_path,
                input_path=mock_chunked_sheet_raw_class_scoped,
                n_books=n_books,
            )

    def test_date_added_to_sheet_names(
        self, mock_chunked_sheet_raw_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the date is added to the sheet names."""
        output_paths = split_chunked_route(
            output_dir=tmp_path,
            input_path=mock_chunked_sheet_raw_class_scoped,
            date=MANIFEST_DATE,
        )
        for output_path in output_paths:
            for sheet_name in pd.ExcelFile(output_path).sheet_names:
                assert str(sheet_name).startswith(f"{MANIFEST_DATE} ")

    def test_sheetname_date_is_friday(
        self, mock_chunked_sheet_raw_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that default date added is Friday."""
        output_paths = split_chunked_route(
            output_dir=tmp_path, input_path=mock_chunked_sheet_raw_class_scoped, n_books=1
        )
        workbook = pd.ExcelFile(output_paths[0])
        this_year = datetime.now().year.__str__()
        for sheet_name in workbook.sheet_names:
            sheet_date = datetime.strptime(
                str(sheet_name).split(" ")[0] + "." + this_year, MANIFEST_DATE_FORMAT + ".%Y"
            )
            assert sheet_date.weekday() == 4

    @pytest.mark.parametrize(
        "output_dir, output_filename, n_books",
        [("", "", 4), ("output", "", 3), ("", "output_filename.xlsx", 1)],
    )
    def test_cli(
        self,
        output_dir: str,
        output_filename: str,
        n_books: int,
        mock_chunked_sheet_raw_class_scoped: Path,
        tmp_path: Path,
    ) -> None:
        """Test CLI works."""
        output_dir = str(tmp_path / output_dir) if output_dir else output_dir
        arg_list = [
            "--input_path",
            str(mock_chunked_sheet_raw_class_scoped),
            "--output_dir",
            output_dir,
            "--output_filename",
            output_filename,
            "--n_books",
            str(n_books),
        ]

        result = subprocess.run(["split_chunked_route"] + arg_list, capture_output=True)
        assert result.returncode == 0

        for i in range(n_books):
            expected_filename = (
                f"{output_filename.split('.')[0]}_{i + 1}.xlsx"
                if output_filename
                else (
                    f"split_workbook_{datetime.now().strftime(FILE_DATE_FORMAT)}"
                    f"_{i + 1}.xlsx"
                )
            )
            expected_output_dir = (
                Path(output_dir) if output_dir else mock_chunked_sheet_raw_class_scoped.parent
            )
            assert (Path(expected_output_dir) / expected_filename).exists()

    def test_output_columns(
        self, mock_chunked_sheet_raw_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the output columns match the SPLIT_ROUTE_COLUMNS constant."""
        output_paths = split_chunked_route(
            output_dir=tmp_path, input_path=mock_chunked_sheet_raw_class_scoped
        )
        for output_path in output_paths:
            workbook = pd.ExcelFile(output_path)
            for sheet_name in workbook.sheet_names:
                driver_sheet = pd.read_excel(workbook, sheet_name=sheet_name)
                assert driver_sheet.columns.to_list() == SPLIT_ROUTE_COLUMNS


class TestCombineRouteTables:
    """combine_route_tables combines driver route CSVs into a single workbook."""

    @pytest.fixture(scope="class")
    def basic_combined_routes(
        self, mock_route_tables_class_scoped: Path, tmp_path_factory: pytest.TempPathFactory
    ) -> Path:
        """Create a basic combined routes table scoped to class for reuse."""
        output_dir = tmp_path_factory.mktemp("tmp_basic_combined_routes", numbered=True)
        output_path = combine_route_tables(
            input_dir=mock_route_tables_class_scoped, output_dir=output_dir
        )
        return output_path

    @pytest.mark.parametrize("output_dir_type", [Path, str])
    @pytest.mark.parametrize("output_dir", ["", "dummy_output"])
    def test_set_output_dir(
        self,
        output_dir_type: type[Path | str],
        output_dir: Path | str,
        tmp_path: Path,
        mock_route_tables_class_scoped: Path,
    ) -> None:
        """Test that the output directory can be set."""
        output_dir = output_dir_type(tmp_path / output_dir)
        output_path = combine_route_tables(
            input_dir=mock_route_tables_class_scoped, output_dir=output_dir
        )
        assert str(output_path.parent) == str(output_dir)

    @pytest.mark.parametrize("output_filename", ["", "dummy_output_filename.xlsx"])
    def test_set_output_filename(
        self, output_filename: str, mock_route_tables_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the output filename can be set."""
        output_path = combine_route_tables(
            output_dir=tmp_path,
            input_dir=mock_route_tables_class_scoped,
            output_filename=output_filename,
        )
        expected_filename = (
            f"combined_routes_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
            if output_filename == ""
            else output_filename
        )
        assert output_path.name == expected_filename

    def test_output_columns(self, basic_combined_routes: Path) -> None:
        """Test that the output columns match the COMBINED_ROUTES_COLUMNS constant."""
        workbook = pd.ExcelFile(basic_combined_routes)
        for sheet_name in workbook.sheet_names:
            driver_sheet = pd.read_excel(workbook, sheet_name=sheet_name)
            assert driver_sheet.columns.to_list() == COMBINED_ROUTES_COLUMNS

    def test_unique_recipients(self, basic_combined_routes: Path) -> None:
        """Test that the recipients don't overlap between the driver route tables.

        By name, address, and phone.
        """
        driver_sheets = _get_driver_sheets(output_paths=[basic_combined_routes])
        combined_output_data = pd.concat(driver_sheets, ignore_index=True)
        assert (
            combined_output_data[[Columns.NAME, Columns.ADDRESS, Columns.PHONE]]
            .duplicated()
            .sum()
            == 0
        )

    def test_complete_contents(
        self, mock_route_tables_class_scoped: Path, basic_combined_routes: Path
    ) -> None:
        """Test that the input data is all covered in the combined workbook."""
        mock_table_paths = list(mock_route_tables_class_scoped.glob("*"))
        full_input_data = pd.concat(
            [pd.read_csv(path) for path in mock_table_paths], ignore_index=True
        ).rename(columns={Columns.PRODUCT_TYPE: Columns.BOX_TYPE})[COMBINED_ROUTES_COLUMNS]
        driver_sheets = _get_driver_sheets(output_paths=[basic_combined_routes])
        combined_output_data = pd.concat(driver_sheets, ignore_index=True)

        full_input_data = full_input_data.sort_values(by=COMBINED_ROUTES_COLUMNS).reset_index(
            drop=True
        )
        combined_output_data = combined_output_data.sort_values(
            by=COMBINED_ROUTES_COLUMNS
        ).reset_index(drop=True)

        pd.testing.assert_frame_equal(full_input_data, combined_output_data)

    @pytest.mark.parametrize("output_dir", ["dummy_output", ""])
    @pytest.mark.parametrize("output_filename", ["", "dummy_output_filename.xlsx"])
    def test_cli(
        self,
        output_dir: str,
        output_filename: str,
        mock_route_tables_class_scoped: Path,
        tmp_path: Path,
    ) -> None:
        """Test CLI works."""
        output_dir = str(tmp_path / output_dir) if output_dir else output_dir
        arg_list = [
            "--input_dir",
            str(mock_route_tables_class_scoped),
            "--output_dir",
            output_dir,
            "--output_filename",
            output_filename,
        ]

        result = subprocess.run(["combine_route_tables"] + arg_list, capture_output=True)
        assert result.returncode == 0

        expected_output_filename = (
            f"combined_routes_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
            if output_filename == ""
            else output_filename
        )
        expected_output_dir = (
            Path(output_dir) if output_dir else mock_route_tables_class_scoped
        )
        assert (expected_output_dir / expected_output_filename).exists()


class TestFormatCombinedRoutes:
    """format_combined_routes formats the combined routes table."""

    @pytest.fixture(scope="class")
    def mock_combined_routes_class_scoped(
        self, tmp_path_factory: pytest.TempPathFactory
    ) -> Path:
        """Mock the combined routes table."""
        tmp_output = tmp_path_factory.mktemp(
            "tmp_mock_combined_routes_class_scoped", numbered=True
        )
        output_path = tmp_output / "combined_routes.xlsx"
        with pd.ExcelWriter(output_path) as writer:
            for driver in DRIVERS:
                df = pd.DataFrame(columns=COMBINED_ROUTES_COLUMNS)
                stops = [stop_no + 1 for stop_no in range(9)]
                df[Columns.STOP_NO] = stops
                df[Columns.NAME] = [f"{driver} Recipient {stop_no}" for stop_no in stops]
                df[Columns.ADDRESS] = [
                    f"{driver} stop {stop_no} address" for stop_no in stops
                ]
                df[Columns.PHONE] = ["13607345215"] * len(stops)
                df[Columns.NOTES] = [f"{driver} stop {stop_no} notes" for stop_no in stops]
                df[Columns.ORDER_COUNT] = [1] * len(stops)
                df[Columns.BOX_TYPE] = [
                    BOX_TYPES[i % len(BOX_TYPES)] for i in range(len(stops))
                ]
                df[Columns.NEIGHBORHOOD] = [
                    NEIGHBORHOODS[i % len(NEIGHBORHOODS)] for i in range(len(stops))
                ]

                assert df.isna().sum().sum() == 0
                assert set(df.columns.to_list()) == set(COMBINED_ROUTES_COLUMNS)

                df.to_excel(writer, sheet_name=f"{MANIFEST_DATE} {driver}", index=False)

        return output_path

    @pytest.fixture(scope="class")
    def mock_combined_routes_ExcelFile_class_scoped(
        self, mock_combined_routes_class_scoped: Path
    ) -> Iterator[pd.ExcelFile]:
        """Mock the combined routes table ExcelFile."""
        with pd.ExcelFile(mock_combined_routes_class_scoped) as xls:
            yield xls

    @pytest.fixture(scope="class")
    def basic_manifest(self, mock_combined_routes_class_scoped: Path) -> Path:
        """Create a basic manifest scoped to class for reuse."""
        output_path = format_combined_routes(input_path=mock_combined_routes_class_scoped)
        return output_path

    @pytest.fixture(scope="class")
    def basic_manifest_workbook(self, basic_manifest: Path) -> Workbook:
        """Create a basic manifest workbook scoped to class for reuse."""
        workbook = load_workbook(basic_manifest)
        return workbook

    @pytest.fixture(scope="class")
    def mock_extra_notes_df_class_scoped(self) -> pd.DataFrame:
        """Mock the extra notes DataFrame."""
        extra_notes_df = pd.DataFrame(
            columns=["tag", "note"],
            data=[
                (
                    "Test extra notes tag 1 *",
                    (
                        "Test extra notes note 1. "
                        "This is a dummy note. It is really long and should be so that we "
                        "can test out column width and word wrapping. It should be long "
                        "enough to wrap around to the next line. And, it should be long "
                        "enough to wrap around to the next line. And, it should be long "
                        "enough to wrap around to the next line. Hopefully, this is long "
                        "enough. Also, hopefully, this is long enough. Further, hopefully, "
                        "this is long enough. Additionally, it will help test out word "
                        "wrapping merged cells."
                    ),
                ),
                ("Test extra notes tag 2 *", "Test extra notes note 2"),
                ("Test extra notes tag 3 *", "Test extra notes note 3"),
            ],
        )
        return extra_notes_df

    @pytest.mark.parametrize("output_dir_type", [Path, str])
    @pytest.mark.parametrize("output_dir", ["", "dummy_output"])
    def test_set_output_dir(
        self,
        output_dir_type: type[Path | str],
        output_dir: Path | str,
        tmp_path: Path,
        mock_combined_routes_class_scoped: Path,
    ) -> None:
        """Test that the output directory can be set."""
        output_dir = output_dir_type(tmp_path / output_dir)
        output_path = format_combined_routes(
            input_path=mock_combined_routes_class_scoped, output_dir=output_dir
        )
        assert str(output_path.parent) == str(output_dir)

    @pytest.mark.parametrize("output_filename", ["", "dummy_output_filename.csv"])
    def test_set_output_filename(
        self, output_filename: str, mock_combined_routes_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the output filename can be set."""
        output_path = format_combined_routes(
            output_dir=tmp_path,
            input_path=mock_combined_routes_class_scoped,
            output_filename=output_filename,
        )
        expected_output_filename = (
            f"formatted_routes_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
            if output_filename == ""
            else output_filename
        )
        assert output_path.name == expected_output_filename

    def test_all_drivers_have_a_sheet(
        self, mock_combined_routes_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that all drivers have a sheet in the formatted workbook. And date works."""
        output_path = format_combined_routes(
            output_dir=tmp_path, input_path=mock_combined_routes_class_scoped
        )
        workbook = pd.ExcelFile(output_path)
        assert set(workbook.sheet_names) == set(
            [f"{MANIFEST_DATE} {driver}" for driver in DRIVERS]
        )

    @pytest.mark.parametrize("output_dir", ["dummy_output", ""])
    @pytest.mark.parametrize("output_filename", ["", "dummy_output_filename.xlsx"])
    def test_cli(
        self,
        output_dir: str,
        output_filename: str,
        mock_combined_routes_class_scoped: Path,
        tmp_path: Path,
    ) -> None:
        """Test CLI works."""
        output_dir = str(tmp_path / output_dir) if output_dir else output_dir
        arg_list = [
            "--input_path",
            str(mock_combined_routes_class_scoped),
            "--output_dir",
            output_dir,
            "--output_filename",
            output_filename,
        ]

        result = subprocess.run(["format_combined_routes"] + arg_list, capture_output=True)
        assert result.returncode == 0

        expected_output_filename = (
            f"formatted_routes_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
            if output_filename == ""
            else output_filename
        )
        expected_output_dir = (
            Path(output_dir) if output_dir else mock_combined_routes_class_scoped.parent
        )
        assert (expected_output_dir / expected_output_filename).exists()

    def test_df_is_same(
        self, mock_combined_routes_ExcelFile_class_scoped: pd.ExcelFile, basic_manifest: Path
    ) -> None:
        """All the input data is in the formatted workbook."""
        for sheet_name in sorted(mock_combined_routes_ExcelFile_class_scoped.sheet_names):
            input_df = pd.read_excel(
                mock_combined_routes_ExcelFile_class_scoped, sheet_name=sheet_name
            )
            input_df.sort_values(by=[Columns.STOP_NO], inplace=True)
            output_df = pd.read_excel(basic_manifest, sheet_name=sheet_name, skiprows=8)

            # Hacky, but need to make sure formatted values haven't fundamentally changed.
            formatted_columns = [Columns.BOX_TYPE, Columns.NAME, Columns.PHONE]
            unformatted_columns = [
                col for col in FORMATTED_ROUTES_COLUMNS if col not in formatted_columns
            ]
            assert input_df[unformatted_columns].equals(output_df[unformatted_columns])

            input_box_type_df = input_df[[Columns.BOX_TYPE]]
            _format_and_validate_box_type(df=input_box_type_df)
            assert input_box_type_df.equals(output_df[[Columns.BOX_TYPE]])

            input_name_df = input_df[[Columns.NAME]]
            _format_and_validate_name(df=input_name_df)
            assert input_name_df.equals(output_df[[Columns.NAME]])

            input_phone_df = input_df[[Columns.PHONE]]
            _format_and_validate_phone(df=input_phone_df)
            assert input_phone_df.equals(output_df[[Columns.PHONE]])

    @pytest.mark.parametrize(
        "cell, expected_value",
        [
            ("A1", "DRIVER SUPPORT: 555-555-5555"),
            ("B1", None),
            ("C1", None),
            ("D1", "RECIPIENT SUPPORT: 555-555-5555 x5"),
            ("E1", None),
            ("F1", "PLEASE SHRED MANIFEST AFTER COMPLETING ROUTE."),
        ],
    )
    def test_header_row(
        self, cell: str, expected_value: str, basic_manifest_workbook: Workbook
    ) -> None:
        """Test that the header row is correct."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws[cell].value == expected_value

    def test_header_row_end(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the header row ends at F1."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            last_non_empty_col = max(
                (cell.column for cell in ws[1] if cell.value), default=None
            )
            assert last_non_empty_col == 6

    @pytest.mark.parametrize("cell", ["A1", "B1", "C1", "D1", "E1", "F1"])
    def test_header_row_color(self, cell: str, basic_manifest_workbook: Workbook) -> None:
        """Test the header row fill color."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws[cell].fill.start_color.rgb == f"{CellColors.HEADER}"

    def test_date_cell(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the date cell is correct."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws["A3"].value == f"Date: {MANIFEST_DATE}"

    def test_driver_cell(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the driver cell is correct."""
        drivers = [driver.upper() for driver in DRIVERS]
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            driver_name = sheet_name.replace(f"{MANIFEST_DATE} ", "")
            assert ws["A5"].value == f"Driver: {driver_name}"
            assert driver_name.upper() in drivers

    def test_agg_cells(
        self,
        mock_combined_routes_ExcelFile_class_scoped: pd.ExcelFile,
        basic_manifest_workbook: Workbook,  # noqa: E501
    ) -> None:
        """Test that the aggregated cells are correct."""
        for sheet_name in sorted(mock_combined_routes_ExcelFile_class_scoped.sheet_names):
            input_df = pd.read_excel(
                mock_combined_routes_ExcelFile_class_scoped, sheet_name=sheet_name
            )
            ws = basic_manifest_workbook[str(sheet_name)]

            agg_dict = _aggregate_route_data(
                df=input_df, extra_notes_df=get_extra_notes(file_path="")
            )

            neighborhoods = ", ".join(agg_dict["neighborhoods"])
            assert ws["A7"].value == f"Neighborhoods: {neighborhoods.upper()}"
            assert ws["E3"].value == BoxType.BASIC
            assert ws["F3"].value == agg_dict["box_counts"][BoxType.BASIC]
            assert ws["E4"].value == BoxType.GF
            assert ws["F4"].value == agg_dict["box_counts"][BoxType.GF]
            assert ws["E5"].value == BoxType.LA
            assert ws["F5"].value == agg_dict["box_counts"][BoxType.LA]
            assert ws["E6"].value == BoxType.VEGAN
            assert ws["F6"].value == agg_dict["box_counts"][BoxType.VEGAN]
            assert ws["E7"].value == "TOTAL BOX COUNT="
            assert ws["F7"].value == agg_dict["total_box_count"]
            assert ws["E8"].value == "PROTEIN COUNT="
            assert ws["F8"].value == agg_dict["protein_box_count"]

    def test_box_type_cell_colors(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the box type cells conditionally formatted with fill color."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            for cell in ws["F"]:
                if cell.row > 9:
                    assert cell.fill.start_color.rgb == f"{BOX_TYPE_COLOR_MAP[cell.value]}"
            for cell in ws["E"]:
                if cell.row > 2 and cell.row < 7:
                    assert cell.fill.start_color.rgb == f"{BOX_TYPE_COLOR_MAP[cell.value]}"

    def test_notes_column_width(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the notes column width is correct."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws["E9"].value == Columns.NOTES
            assert ws.column_dimensions["E"].width == NOTES_COLUMN_WIDTH

    @pytest.mark.parametrize(
        "cell",
        [
            # Header row.
            "A1",
            "B1",
            "C1",
            "D1",
            "E1",
            "F1",
            # Aggregated data.
            "A3",
            "A5",
            "A7",
            "E3",
            "E4",
            "E5",
            "E6",
            "E7",
            "E8",
            "F3",
            "F4",
            "F5",
            "F6",
            "F7",
            "F8",
            # Data header.
            "A9",
            "B9",
            "C9",
            "D9",
            "E9",
            "F9",
        ],
    )
    def test_bold_cells(self, cell: str, basic_manifest_workbook: Workbook) -> None:
        """Test that the cells are bold."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws[cell].font.bold

    def test_cell_right_alignment(self, basic_manifest_workbook: Workbook) -> None:
        """Test right-aligned cells."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            right_aligned_cells = [ws["D1"], ws["F1"]] + [
                cell for row in ws["E3:F8"] for cell in row
            ]
            for cell in right_aligned_cells:
                assert cell.alignment.horizontal == "right"

    def test_cell_left_alignment(self, basic_manifest_workbook: Workbook) -> None:
        """Test left-aligned cells."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            left_aligned_cells = [cell for row in ws["A1:A8"] for cell in row] + [
                cell for row in ws["A9:F9"] for cell in row
            ]
            for cell in left_aligned_cells:
                assert cell.alignment.horizontal == "left"

    @pytest.mark.parametrize("extra_notes_file", ["", "dummy_extra_notes.csv"])
    def test_extra_notes(
        self,
        extra_notes_file: str,
        mock_combined_routes_class_scoped: Path,
        mock_extra_notes_df_class_scoped: pd.DataFrame,
        tmp_path: Path,
    ) -> None:
        """Test that extra notes are added to the manifest."""
        mock_extra_notes_context, extra_notes_file = _get_extra_notes(
            extra_notes_file=extra_notes_file,
            extra_notes_dir=str(mock_combined_routes_class_scoped.parent),
            extra_notes_df=mock_extra_notes_df_class_scoped,
        )

        new_mock_combined_routes_path = (
            mock_combined_routes_class_scoped.parent / "new_mock_combined_routes.xlsx"
        )
        mock_combined_routes_file = pd.ExcelFile(mock_combined_routes_class_scoped)

        first_sheet_name = str(mock_combined_routes_file.sheet_names[0])
        first_df = mock_combined_routes_file.parse(sheet_name=first_sheet_name)
        second_sheet_name = str(mock_combined_routes_file.sheet_names[1])
        second_df = mock_combined_routes_file.parse(sheet_name=second_sheet_name)
        first_df, second_df = _set_extra_notes(
            first_df=first_df,
            second_df=second_df,
            extra_notes_df=mock_extra_notes_df_class_scoped,
        )

        with pd.ExcelWriter(new_mock_combined_routes_path) as writer:
            first_df.to_excel(writer, sheet_name=first_sheet_name, index=False)
            second_df.to_excel(writer, sheet_name=second_sheet_name, index=False)
            for sheet_name in mock_combined_routes_file.sheet_names[2:]:
                df = mock_combined_routes_file.parse(sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=str(sheet_name), index=False)

        with mock_extra_notes_context:
            manifests_path = format_combined_routes(
                output_dir=tmp_path,
                input_path=new_mock_combined_routes_path,
                extra_notes_file=extra_notes_file,
            )

        _assert_extra_notes(
            manifests_path=manifests_path,
            first_sheet_name=first_sheet_name,
            second_sheet_name=second_sheet_name,
            extra_notes_df=mock_extra_notes_df_class_scoped,
            first_df=first_df,
            second_df=second_df,
        )


class TestCreateManifestsClassScoped:
    """create_manifests formats the route tables CSVs."""

    @pytest.fixture(scope="class")
    def basic_manifest(
        self, mock_route_tables_class_scoped: Path, tmp_path_factory: pytest.TempPathFactory
    ) -> Path:
        """Create a basic manifest scoped to class for reuse."""
        output_dir = tmp_path_factory.mktemp("tmp_basic_manifest", numbered=True)
        output_path = create_manifests(
            input_dir=mock_route_tables_class_scoped, output_dir=output_dir
        )
        return output_path

    @pytest.fixture(scope="class")
    def basic_manifest_workbook(self, basic_manifest: Path) -> Workbook:
        """Create a basic manifest workbook scoped to class for reuse."""
        workbook = load_workbook(basic_manifest)
        return workbook

    @pytest.fixture(scope="class")
    def basic_manifest_ExcelFile(self, basic_manifest: Path) -> Iterator[pd.ExcelFile]:
        """Create a basic manifest workbook scoped to class for reuse."""
        with pd.ExcelFile(basic_manifest) as xls:
            yield xls

    @pytest.mark.parametrize("output_dir_type", [Path, str])
    # TODO: Empty string here doesn't really test anything. Mock os.getcwd?
    @pytest.mark.parametrize("output_dir", ["", "dummy_output"])
    def test_set_output_dir(
        self,
        output_dir_type: type[Path | str],
        output_dir: Path | str,
        mock_route_tables_class_scoped: Path,
        tmp_path: Path,
    ) -> None:
        """Test that the output directory can be set."""
        output_dir = output_dir_type(tmp_path / output_dir)
        output_path = create_manifests(
            input_dir=mock_route_tables_class_scoped, output_dir=output_dir
        )
        assert str(output_path.parent) == str(output_dir)

    @pytest.mark.parametrize("output_filename", ["", "dummy_output_filename.csv"])
    def test_set_output_filename(
        self, output_filename: str, mock_route_tables_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the output filename can be set."""
        output_path = create_manifests(
            output_dir=tmp_path,
            input_dir=mock_route_tables_class_scoped,
            output_filename=output_filename,
        )
        expected_output_filename = (
            f"final_manifests_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
            if output_filename == ""
            else output_filename
        )
        assert output_path.name == expected_output_filename

    # TODO: Create basic fixtures for this call and after.
    def test_all_drivers_have_a_sheet(
        self, mock_route_tables_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that all drivers have a sheet in the formatted workbook. And date works."""
        output_path = create_manifests(
            output_dir=tmp_path, input_dir=mock_route_tables_class_scoped
        )
        workbook = pd.ExcelFile(output_path)
        assert set(workbook.sheet_names) == set(
            [f"{MANIFEST_DATE} {driver}" for driver in DRIVERS]
        )

    def test_date_field_matches_sheet_date(
        self, mock_route_tables_class_scoped: Path, tmp_path: Path
    ) -> None:
        """Test that the date field matches the sheet date."""
        output_path = create_manifests(
            input_dir=mock_route_tables_class_scoped, output_dir=tmp_path
        )
        workbook = load_workbook(output_path)
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            field_date = ws["A3"].value.split(" ")[1]
            sheet_name_date = sheet_name.split(" ")[0]
            assert field_date == sheet_name_date

    @pytest.mark.parametrize("output_dir", ["dummy_output", ""])
    @pytest.mark.parametrize("output_filename", ["", "dummy_output_filename.xlsx"])
    def test_cli(
        self,
        output_dir: str,
        output_filename: str,
        mock_route_tables_class_scoped: Path,
        tmp_path: Path,
    ) -> None:
        """Test CLI works."""
        output_dir = str(tmp_path / output_dir) if output_dir else output_dir
        arg_list = [
            "--input_dir",
            str(mock_route_tables_class_scoped),
            "--output_dir",
            output_dir,
            "--output_filename",
            output_filename,
        ]

        result = subprocess.run(["create_manifests"] + arg_list, capture_output=True)
        assert result.returncode == 0

        expected_output_filename = (
            f"final_manifests_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
            if output_filename == ""
            else output_filename
        )
        expected_output_dir = (
            Path(output_dir) if output_dir else mock_route_tables_class_scoped
        )
        assert (expected_output_dir / expected_output_filename).exists()

    def test_df_is_same(
        self, mock_route_tables_class_scoped: Path, basic_manifest_ExcelFile: pd.ExcelFile
    ) -> None:
        """All the input data is in the formatted workbook."""
        for sheet_name in sorted(basic_manifest_ExcelFile.sheet_names):
            input_df = pd.read_csv(mock_route_tables_class_scoped / f"{sheet_name}.csv")
            output_df = pd.read_excel(
                basic_manifest_ExcelFile, sheet_name=sheet_name, skiprows=8
            )

            # Hacky, but need to make sure formatted values haven't fundamentally changed.
            formatted_columns = [Columns.BOX_TYPE, Columns.NAME, Columns.PHONE]
            unformatted_columns = [
                col for col in FORMATTED_ROUTES_COLUMNS if col not in formatted_columns
            ]
            assert input_df[unformatted_columns].equals(output_df[unformatted_columns])

            input_df.rename(columns={Columns.PRODUCT_TYPE: Columns.BOX_TYPE}, inplace=True)
            input_box_type_df = input_df[[Columns.BOX_TYPE]]
            _format_and_validate_box_type(df=input_box_type_df)
            assert input_box_type_df.equals(output_df[[Columns.BOX_TYPE]])

            input_name_df = input_df[[Columns.NAME]]
            _format_and_validate_name(df=input_name_df)
            assert input_name_df.equals(output_df[[Columns.NAME]])

            input_phone_df = input_df[[Columns.PHONE]]
            _format_and_validate_phone(df=input_phone_df)
            assert input_phone_df.equals(output_df[[Columns.PHONE]])


# TODO: Revisit moving the rest to class scope once output dirs cconsolidated.
# Conflicts now.
class TestCreateManifests:
    """create_manifests formats the route tables CSVs."""

    @pytest.fixture()
    def basic_manifest(self, mock_route_tables: Path) -> Path:
        """Create a basic manifest scoped to class for reuse."""
        output_path = create_manifests(input_dir=mock_route_tables)
        return output_path

    @pytest.fixture()
    def basic_manifest_workbook(self, basic_manifest: Path) -> Workbook:
        """Create a basic manifest workbook scoped to class for reuse."""
        workbook = load_workbook(basic_manifest)
        return workbook

    @pytest.fixture()
    def basic_manifest_ExcelFile(self, basic_manifest: Path) -> Iterator[pd.ExcelFile]:
        """Create a basic manifest workbook scoped to class for reuse."""
        with pd.ExcelFile(basic_manifest) as xls:
            yield xls

    @pytest.mark.parametrize(
        "cell, expected_value",
        [
            ("A1", "DRIVER SUPPORT: 555-555-5555"),
            ("B1", None),
            ("C1", None),
            ("D1", "RECIPIENT SUPPORT: 555-555-5555 x5"),
            ("E1", None),
            ("F1", "PLEASE SHRED MANIFEST AFTER COMPLETING ROUTE."),
        ],
    )
    def test_header_row(
        self, cell: str, expected_value: str, basic_manifest_workbook: Workbook
    ) -> None:
        """Test that the header row is correct."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws[cell].value == expected_value

    def test_header_row_end(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the header row ends at F1."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            last_non_empty_col = max(
                (cell.column for cell in ws[1] if cell.value), default=None
            )
            assert last_non_empty_col == 6

    @pytest.mark.parametrize("cell", ["A1", "B1", "C1", "D1", "E1", "F1"])
    def test_header_row_color(self, cell: str, basic_manifest_workbook: Workbook) -> None:
        """Test the header row fill color."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws[cell].fill.start_color.rgb == f"{CellColors.HEADER}"

    def test_date_cell(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the date cell is correct."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws["A3"].value == f"Date: {MANIFEST_DATE}"

    def test_driver_cell(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the driver cell is correct."""
        drivers = [driver.upper() for driver in DRIVERS]
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            driver_name = sheet_name.replace(f"{MANIFEST_DATE} ", "")
            assert ws["A5"].value == f"Driver: {driver_name}"
            assert driver_name.upper() in drivers

    def test_agg_cells(
        self, basic_manifest_workbook: Workbook, mock_route_tables: Path
    ) -> None:
        """Test that the aggregated cells are correct."""
        for sheet_name in sorted(basic_manifest_workbook.sheetnames):
            input_df = pd.read_csv(mock_route_tables / f"{sheet_name}.csv")
            ws = basic_manifest_workbook[sheet_name]

            input_df.rename(columns={Columns.PRODUCT_TYPE: Columns.BOX_TYPE}, inplace=True)
            agg_dict = _aggregate_route_data(
                df=input_df, extra_notes_df=get_extra_notes(file_path="")
            )

            neighborhoods = ", ".join(agg_dict["neighborhoods"])
            assert ws["A7"].value == f"Neighborhoods: {neighborhoods.upper()}"
            assert ws["E3"].value == BoxType.BASIC
            assert ws["F3"].value == agg_dict["box_counts"][BoxType.BASIC]
            assert ws["E4"].value == BoxType.GF
            assert ws["F4"].value == agg_dict["box_counts"][BoxType.GF]
            assert ws["E5"].value == BoxType.LA
            assert ws["F5"].value == agg_dict["box_counts"][BoxType.LA]
            assert ws["E6"].value == BoxType.VEGAN
            assert ws["F6"].value == agg_dict["box_counts"][BoxType.VEGAN]
            assert ws["E7"].value == "TOTAL BOX COUNT="
            assert ws["F7"].value == agg_dict["total_box_count"]
            assert ws["E8"].value == "PROTEIN COUNT="
            assert ws["F8"].value == agg_dict["protein_box_count"]

    def test_box_type_cell_colors(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the box type cells conditionally formatted with fill color."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            for cell in ws["F"]:
                if cell.row > 9:
                    assert cell.fill.start_color.rgb == f"{BOX_TYPE_COLOR_MAP[cell.value]}"
            for cell in ws["E"]:
                if cell.row > 2 and cell.row < 7:
                    assert cell.fill.start_color.rgb == f"{BOX_TYPE_COLOR_MAP[cell.value]}"

    def test_notes_column_width(self, basic_manifest_workbook: Workbook) -> None:
        """Test that the notes column width is correct."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws["E9"].value == Columns.NOTES
            assert ws.column_dimensions["E"].width == NOTES_COLUMN_WIDTH

    @pytest.mark.parametrize(
        "cell",
        [
            # Header row.
            "A1",
            "B1",
            "C1",
            "D1",
            "E1",
            "F1",
            # Aggregated data.
            "A3",
            "A5",
            "A7",
            "E3",
            "E4",
            "E5",
            "E6",
            "E7",
            "E8",
            "F3",
            "F4",
            "F5",
            "F6",
            "F7",
            "F8",
            # Data header.
            "A9",
            "B9",
            "C9",
            "D9",
            "E9",
            "F9",
        ],
    )
    def test_bold_cells(self, cell: str, basic_manifest_workbook: Workbook) -> None:
        """Test that the cells are bold."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            assert ws[cell].font.bold

    def test_cell_right_alignment(self, basic_manifest_workbook: Workbook) -> None:
        """Test right-aligned cells."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            right_aligned_cells = [ws["D1"], ws["F1"]] + [
                cell for row in ws["E3:F8"] for cell in row
            ]
            for cell in right_aligned_cells:
                assert cell.alignment.horizontal == "right"

    def test_cell_left_alignment(self, basic_manifest_workbook: Workbook) -> None:
        """Test left-aligned cells."""
        for sheet_name in basic_manifest_workbook.sheetnames:
            ws = basic_manifest_workbook[sheet_name]
            left_aligned_cells = [cell for row in ws["A1:A8"] for cell in row] + [
                cell for row in ws["A9:F9"] for cell in row
            ]
            for cell in left_aligned_cells:
                assert cell.alignment.horizontal == "left"

    @pytest.mark.parametrize("extra_notes_file", ["", "dummy_extra_notes.csv"])
    def test_extra_notes(
        self,
        extra_notes_file: str,
        mock_route_tables: Path,
        mock_extra_notes_df: pd.DataFrame,
    ) -> None:
        """Test that extra notes are added to the manifest."""
        mock_extra_notes_context, extra_notes_file = _get_extra_notes(
            extra_notes_file=extra_notes_file,
            extra_notes_dir=str(mock_route_tables.parent),
            extra_notes_df=mock_extra_notes_df,
        )

        mock_route_tables_names = glob.glob(str(mock_route_tables / "*.csv"))
        first_sheet_name = Path(mock_route_tables_names[0]).stem
        first_df = pd.read_csv(mock_route_tables_names[0])
        first_df = pd.concat([first_df] * 5, ignore_index=True)
        first_df[Columns.STOP_NO] = range(1, len(first_df) + 1)
        second_sheet_name = Path(mock_route_tables_names[1]).stem
        second_df = pd.read_csv(mock_route_tables_names[1])
        first_df, second_df = _set_extra_notes(
            first_df=first_df, second_df=second_df, extra_notes_df=mock_extra_notes_df
        )
        first_df.to_csv(mock_route_tables_names[0], index=False)
        second_df.to_csv(mock_route_tables_names[1], index=False)

        with mock_extra_notes_context:
            manifests_path = create_manifests(
                input_dir=mock_route_tables, extra_notes_file=extra_notes_file
            )

        _assert_extra_notes(
            manifests_path=manifests_path,
            first_sheet_name=first_sheet_name,
            second_sheet_name=second_sheet_name,
            extra_notes_df=mock_extra_notes_df,
            first_df=first_df,
            second_df=second_df,
        )


@pytest.mark.parametrize(
    "route_df, extra_notes_df, expected_agg_dict, error_context",
    [
        (
            pd.DataFrame(
                {
                    Columns.BOX_TYPE: ["BASIC", "GF", "LA", "BASIC", "GF", "LA", "Vegan"],
                    Columns.ORDER_COUNT: [1, 1, 1, 2, 1, 1, 2],
                    Columns.NEIGHBORHOOD: [
                        "YORK",
                        "YORK",
                        "YORK",
                        "PUGET",
                        "YORK",
                        "YORK",
                        "PUGET",
                    ],
                    Columns.NOTES: ["", "", "", "", "", "", ""],
                }
            ),
            pd.DataFrame(columns=["tag", "note"]),
            {
                "box_counts": {"BASIC": 3, "GF": 2, "LA": 2, "VEGAN": 2},
                "total_box_count": 9,
                "protein_box_count": 7,
                "neighborhoods": ["YORK", "PUGET"],
                "extra_notes": [],
            },
            nullcontext(),
        ),
        (
            pd.DataFrame(
                {
                    Columns.BOX_TYPE: ["BASIC", "GF", "LA", "BASIC", "GF", "LA"],
                    Columns.ORDER_COUNT: [1, 1, 1, 2, 1, 1],
                    Columns.NEIGHBORHOOD: ["YORK", "YORK", "YORK", "PUGET", "YORK", "YORK"],
                    Columns.NOTES: ["Test tag * asfgasfg", "", "", "", "", ""],
                }
            ),
            pd.DataFrame(columns=["tag", "note"], data=[("Test tag *", "Test note.")]),
            {
                "box_counts": {"BASIC": 3, "GF": 2, "LA": 2, "VEGAN": 0},
                "total_box_count": 7,
                "protein_box_count": 7,
                "neighborhoods": ["YORK", "PUGET"],
                "extra_notes": ["* Test tag: Test note."],
            },
            nullcontext(),
        ),
        (
            pd.DataFrame(
                {
                    Columns.BOX_TYPE: [
                        "BASIC",
                        "GF",
                        "LA",
                        "BASIC",
                        "GF",
                        "LA",
                        "Vegan",
                        "bad box type",
                    ],
                    Columns.ORDER_COUNT: [1, 1, 1, 2, 1, 1, 2, 1],
                    Columns.NEIGHBORHOOD: [
                        "YORK",
                        "YORK",
                        "YORK",
                        "PUGET",
                        "YORK",
                        "YORK",
                        "PUGET",
                        "PUGET",
                    ],
                }
            ),
            pd.DataFrame(columns=["tag", "note"]),
            {},
            pytest.raises(
                ValueError,
                match=re.escape("Invalid box type in route data: {'BAD BOX TYPE'}"),
            ),
        ),
    ],
)
def test_aggregate_route_data(
    route_df: pd.DataFrame,
    extra_notes_df: pd.DataFrame,
    expected_agg_dict: dict,
    error_context: AbstractContextManager,
) -> None:
    """Test that a route's data is aggregated correctly."""
    with error_context:
        agg_dict = _aggregate_route_data(df=route_df, extra_notes_df=extra_notes_df)
        assert agg_dict == expected_agg_dict


@pytest.mark.parametrize(
    "driver_sets, expected_driver_sets",
    [
        (
            [["Driver A", "Driver B"], ["Driver C", "Driver D"]],
            [["Driver A", "Driver B"], ["Driver C", "Driver D"]],
        ),
        (
            [["Driver A", "Driver B #1"], ["Driver B #2", "Driver C"]],
            [["Driver A", "Driver B #1", "Driver B #2"], ["Driver C"]],
        ),
        (
            [["Driver A #1", "Driver B"], ["Driver A #2", "Driver C"]],
            [["Driver A #1", "Driver B", "Driver A #2"], ["Driver C"]],
        ),
        (
            [
                ["Driver A", "Driver B #1"],
                ["Driver C", "Driver D"],
                ["Driver B #2", "Driver E"],
            ],
            [
                ["Driver A", "Driver B #1", "Driver B #2"],
                ["Driver C", "Driver D"],
                ["Driver E"],
            ],
        ),
    ],
)
def test_group_numbered_drivers(
    driver_sets: list[list[str]], expected_driver_sets: list[list[str]]
) -> None:
    """Test that numbered drivers are grouped correctly."""
    returned_driver_sets = _group_numbered_drivers(driver_sets=driver_sets)
    assert sorted(returned_driver_sets) == sorted(expected_driver_sets)


@pytest.mark.parametrize(
    "drivers, n_books, expected_driver_sets",
    [
        (
            ["Driver A", "Driver B", "Driver C", "Driver D"],
            2,
            [["Driver A", "Driver B"], ["Driver C", "Driver D"]],
        ),
        (
            ["Driver A", "Driver B #1", "Driver B #2", "Driver C"],
            2,
            [["Driver A", "Driver B #1", "Driver B #2"], ["Driver C"]],
        ),
        (
            ["Driver A", "Driver B #1", "Driver C", "Driver B #2"],
            2,
            [["Driver A", "Driver B #1", "Driver B #2"], ["Driver C"]],
        ),
        (
            ["Driver A", "Driver B #1", "Driver C", "Driver D", "Driver B #2", "Driver E"],
            3,
            [
                ["Driver A", "Driver B #1", "Driver B #2"],
                ["Driver C"],
                ["Driver D", "Driver E"],
            ],
        ),
    ],
)
def test_get_driver_sets_group_numbered(
    drivers: list[str], n_books: int, expected_driver_sets: list[list[str]]
) -> None:
    """Test that numbered drivers are grouped correctly."""
    returned_driver_sets = _get_driver_sets(
        drivers=drivers, n_books=n_books, book_one_drivers_file=""
    )
    assert returned_driver_sets == expected_driver_sets


def _get_extra_notes(
    extra_notes_file: str, extra_notes_dir: str, extra_notes_df: pd.DataFrame
) -> tuple[AbstractContextManager, str]:
    mock_extra_notes_context = nullcontext()
    if extra_notes_file:
        extra_notes_file = f"{extra_notes_dir}/{extra_notes_file}"
        extra_notes_df.to_csv(extra_notes_file, index=False)
    else:

        class TestExtraNotes:
            df: Final[pd.DataFrame] = extra_notes_df

        mock_extra_notes_context = patch(
            "bfb_delivery.lib.formatting.utils.ExtraNotes", new=TestExtraNotes
        )

    return mock_extra_notes_context, extra_notes_file


def _set_extra_notes(
    first_df: pd.DataFrame, second_df: pd.DataFrame, extra_notes_df: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    first_df[Columns.NOTES] = [
        extra_notes_df["tag"].iloc[0],
        extra_notes_df["tag"].iloc[1],
        extra_notes_df["tag"].iloc[1],
    ] + first_df[Columns.NOTES].to_list()[3:]

    second_df[Columns.NOTES] = [extra_notes_df["tag"].iloc[2]] + second_df[
        Columns.NOTES
    ].to_list()[1:]

    return first_df, second_df


def _assert_extra_notes(
    manifests_path: Path,
    first_sheet_name: str,
    second_sheet_name: str,
    extra_notes_df: pd.DataFrame,
    first_df: pd.DataFrame,
    second_df: pd.DataFrame,
) -> None:
    manifests_workbook = load_workbook(manifests_path)
    first_ws = manifests_workbook[first_sheet_name]
    second_ws = manifests_workbook[second_sheet_name]
    start_first_notes = 11 + len(first_df)
    start_second_notes = 11 + len(second_df)

    assert first_ws[f"E{start_first_notes}"].value.startswith(
        "* " + extra_notes_df["tag"].iloc[0].replace("*", "").strip() + ": "
    )
    assert extra_notes_df["note"].iloc[0] in first_ws[f"E{start_first_notes}"].value
    assert first_ws[f"E{start_first_notes + 1}"].value.startswith(
        "* " + extra_notes_df["tag"].iloc[1].replace("*", "").strip() + ": "
    )
    assert extra_notes_df["note"].iloc[1] in first_ws[f"E{start_first_notes + 1}"].value
    assert second_ws[f"E{start_second_notes}"].value.startswith(
        "* " + extra_notes_df["tag"].iloc[2].replace("*", "").strip() + ": "
    )
    assert extra_notes_df["note"].iloc[2] in second_ws[f"E{start_second_notes}"].value

    return


def _get_driver_sheets(output_paths: list[Path]) -> list[pd.DataFrame]:
    driver_sheets = []
    for output_path in output_paths:
        workbook = pd.ExcelFile(output_path)
        driver_sheets = driver_sheets + [
            pd.read_excel(workbook, sheet_name=sheet) for sheet in workbook.sheet_names
        ]

    return driver_sheets
